import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

# Define output path
OUTPUT_DIR = r"C:\Users\prajw\OneDrive\Desktop\Albert\Albert_Project\Heritage\03_Shop_Drawings"
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "05_Cabinet_Estimation_Shop_Drawings_Heritage_Village.xlsx")

wb = openpyxl.Workbook()

# Setup sheets
ws_info = wb.active
ws_info.title = "Project Info"
ws_matrix = wb.create_sheet("Cabinet Matrix")
ws_library = wb.create_sheet("Cabinet Library")
ws_costing = wb.create_sheet("Job Costing")

# Standard styling
font_title = Font(name="Arial", size=16, bold=True, color="1B365D")
font_section = Font(name="Arial", size=12, bold=True, color="1B365D")
font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Arial", size=10, bold=True)
font_regular = Font(name="Arial", size=10)
font_small = Font(name="Arial", size=8, italic=True)

fill_navy = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
fill_light_gray = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
fill_blue_input = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
fill_green_calc = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

thin_border_side = Side(style="thin", color="D3D3D3")
thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
double_bottom = Border(bottom=Side(style="double", color="1B365D"), top=thin_border_side)

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

# -------------------------------------------------------------
# SHEET 1: Project Info
# -------------------------------------------------------------
ws_info.views.sheetView[0].showGridLines = True
ws_info.cell(1, 1, "CABINET ESTIMATION — PROJECT INFORMATION").font = font_title
ws_info.cell(2, 1, "Heritage Village — Atlantic Pacific Communities").font = font_small

info_fields = [
    ("Project Name:", "Heritage Village", fill_blue_input),
    ("Project ID:", "23-045", fill_blue_input),
    ("Date:", "=TODAY()", None),
    ("Prepared By:", "AI Estimation System", None),
    ("Building Type:", "Residential — Type V-A", None),
    ("Price List Used:", "MS PRICE LIST LEVEL 1 -90CM", fill_blue_input),
    ("GP Target %:", 0.35, fill_blue_input),
    ("Total Units:", "=SUM('Cabinet Matrix'!B6:B19)", fill_green_calc),
]

row_idx = 4
for label, val, fill in info_fields:
    ws_info.cell(row_idx, 1, label).font = font_bold
    cell = ws_info.cell(row_idx, 2, val)
    cell.font = font_regular
    if fill:
        cell.fill = fill
    if label == "GP Target %:":
        cell.number_format = "0.0%"
    row_idx += 1

# -------------------------------------------------------------
# SHEET 2: Cabinet Library
# -------------------------------------------------------------
ws_library.views.sheetView[0].showGridLines = True
ws_library.cell(1, 1, "CABINET LIBRARY — STANDARD SPECIFICATIONS").font = font_title

lib_headers = ["Cabinet Code", "Description", "Type", "Width (in)", "Height (in)", "Depth (in)", "Finish Tier", "Door Style", "Notes"]
for col, h in enumerate(lib_headers, 1):
    cell = ws_library.cell(3, col, h)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = align_center

lib_data = [
    ("W3012", "Wall Cabinet 30W x 12H (Fridge)", "Upper", 30, 12, 12, "Standard 1", "Shaker", "Over fridge"),
    ("W3618", "Wall Cabinet 36W x 18H (Microwave)", "Upper", 36, 18, 12, "Standard 1", "Shaker", "Over microwave"),
    ("W3630", "Wall Cabinet 36W x 30H", "Upper", 36, 30, 12, "Standard 1", "Shaker", "Standard upper"),
    ("W3030", "Wall Cabinet 30W x 30H", "Upper", 30, 30, 12, "Standard 1", "Shaker", "Standard upper"),
    ("WC2430", "Corner Wall Cabinet 24W", "Upper", 24, 30, 12, "Standard 1", "Shaker", "Blind corner"),
    ("B36", "Base Cabinet 36W", "Lower", 36, 34.5, 24, "Standard 1", "Shaker", "Standard base"),
    ("B30", "Base Cabinet 30W", "Lower", 30, 34.5, 24, "Standard 1", "Shaker", "Standard base"),
    ("B24", "Base Cabinet 24W", "Lower", 24, 34.5, 24, "Standard 1", "Shaker", "Standard base"),
    ("B18", "Base Cabinet 18W", "Lower", 18, 34.5, 24, "Standard 1", "Shaker", "Standard base"),
    ("BC36", "Corner Base Cabinet", "Lower", 36, 34.5, 24, "Standard 1", "Shaker", "Lazy susan option"),
    ("T1884", "Tall Pantry 18W x 84H", "Tall", 18, 84, 24, "Standard 1", "Shaker", "Pantry"),
    ("T2484", "Tall Pantry 24W x 84H", "Tall", 24, 84, 24, "Standard 1", "Shaker", "Pantry"),
    ("T1584", "Tall Pantry 15W x 84H", "Tall", 15, 84, 24, "Standard 1", "Shaker", "Pantry (Heritage Spec)"),
    ("B-RO", "Base Cabinet with Roll-out Drawers", "Lower", 36, 34.5, 24, "Standard 1", "Shaker", "Roll-out drawers"),
    ("VAN24", "Bath Vanity 24W", "Vanity", 24, 34.5, 21, "Standard 1", "Shaker", "Single sink"),
    ("VAN30", "Bath Vanity 30W", "Vanity", 30, 34.5, 21, "Standard 1", "Shaker", "Single sink"),
    ("VAN36", "Bath Vanity 36W", "Vanity", 36, 34.5, 21, "Standard 1", "Shaker", "Single sink"),
    ("VAN48", "Bath Vanity 48W", "Vanity", 48, 34.5, 21, "Standard 1", "Shaker", "Double sink"),
    ("VAN60", "Bath Vanity 60W", "Vanity", 60, 34.5, 21, "Standard 1", "Shaker", "Double sink"),
    ("MED24", "Medicine Cabinet 24W", "Bath", 24, 30, 4, "Standard 1", "Mirror", "Surface mount"),
    ("MED30", "Medicine Cabinet 30W", "Bath", 30, 30, 4, "Standard 1", "Mirror", "Surface mount"),
    ("MED36", "Medicine Cabinet 36W", "Bath", 36, 30, 4, "Standard 1", "Mirror", "Surface mount"),
    ("MED60", "Medicine Cabinet 60W", "Bath", 60, 30, 4, "Standard 1", "Mirror", "Surface mount"),
    ("LIN18", "Linen Cabinet 18W x 84H", "Bath", 18, 84, 18, "Standard 1", "Shaker", "Linen"),
]

for row_offset, row in enumerate(lib_data, 4):
    for col_idx, val in enumerate(row, 1):
        cell = ws_library.cell(row_offset, col_idx, val)
        cell.font = font_regular
        cell.border = thin_border
        if col_idx in [4, 5, 6]:
            cell.alignment = align_right
        else:
            cell.alignment = align_left

# -------------------------------------------------------------
# SHEET 3: Cabinet Matrix
# -------------------------------------------------------------
ws_matrix.views.sheetView[0].showGridLines = True
ws_matrix.cell(1, 1, "CABINET MATRIX — UNIT TYPE BREAKDOWN").font = font_title

headers_m1 = ["Unit Type", "Qty", "Floor(s)", "KITCHEN — Upper Cabinets", "", "", "", "KITCHEN — Lower / Base Cabinets", "", "", "", "Kitchen Totals", "", "BATH — Vanities", "", "", "Bath Total", "Unit Grand Total"]
headers_m2 = ["Unit Type", "Qty", "Floor(s)", "Wall Cab", "Wall Cab", "Corner Wall", "Specialty", "Base Cab", "Base Cab", "Corner Base", "Specialty", "Kitchen Sub", "Kitchen Total", "Single Van", "Double Van", "Linen/Med", "Bath Sub", "UNIT TOTAL"]

for col, h in enumerate(headers_m1, 1):
    cell = ws_matrix.cell(4, col, h)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = align_center

for col, h in enumerate(headers_m2, 1):
    cell = ws_matrix.cell(5, col, h)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = align_center

# Merge columns in headers
ws_matrix.merge_cells("A4:A5")
ws_matrix.merge_cells("B4:B5")
ws_matrix.merge_cells("C4:C5")
ws_matrix.merge_cells("D4:G4")
ws_matrix.merge_cells("H4:K4")
ws_matrix.merge_cells("L4:M4")
ws_matrix.merge_cells("N4:P4")
ws_matrix.merge_cells("Q4:Q5")
ws_matrix.merge_cells("R4:R5")

# 14 Heritage Unit types
matrix_data = [
    ("A-1", 15, "Ground/2nd/3rd", 3, 0, 1, 1, 4, 0, 1, 1, 1, 0, 2),
    ("A-1a ACC", 3, "Ground/2nd/3rd", 3, 0, 1, 1, 4, 0, 1, 1, 1, 0, 2),
    ("B-1", 12, "Ground/2nd/3rd", 4, 0, 1, 1, 5, 0, 1, 1, 2, 0, 3),
    ("B-1a ACC", 3, "Ground/2nd/3rd", 4, 0, 1, 1, 5, 0, 1, 1, 2, 0, 3),
    ("C-1", 8, "Ground/2nd/3rd", 4, 0, 1, 1, 5, 0, 1, 1, 2, 0, 3),
    ("C-2", 6, "Ground/2nd/3rd", 4, 0, 1, 1, 5, 0, 1, 1, 2, 0, 3),
    ("C-2a ACC", 2, "Ground/2nd/3rd", 4, 0, 1, 1, 5, 0, 1, 1, 2, 0, 3),
    ("C-3", 6, "Ground/2nd/3rd", 4, 0, 1, 1, 5, 0, 1, 1, 2, 0, 3),
    ("D-1N", 4, "Ground/2nd/3rd", 5, 0, 1, 1, 6, 0, 1, 1, 2, 0, 3),
    ("D-1", 1, "Ground/2nd/3rd", 5, 0, 1, 1, 6, 0, 1, 1, 2, 0, 3),
    ("D-1a ACC", 2, "Ground/2nd/3rd", 5, 0, 1, 1, 6, 0, 1, 1, 2, 0, 3),
    ("D-1A ACC", 2, "Ground/2nd/3rd", 5, 0, 1, 1, 6, 0, 1, 1, 2, 0, 3),
    ("ST-1a ACC", 4, "Ground/2nd/3rd", 2, 0, 0, 1, 3, 0, 0, 0, 1, 0, 1),
    ("ST-1", 6, "Ground/2nd/3rd", 2, 0, 0, 1, 3, 0, 0, 0, 1, 0, 1),
]

for row_offset, row in enumerate(matrix_data, 6):
    ws_matrix.cell(row_offset, 1, row[0]).font = font_bold
    ws_matrix.cell(row_offset, 2, row[1]).font = font_regular
    ws_matrix.cell(row_offset, 3, row[2]).font = font_regular
    
    # counts
    for c in range(3, 11):
        ws_matrix.cell(row_offset, c + 1, row[c]).font = font_regular
    
    # Formulas
    # Kitchen Sub = Sum of Upper (D:G) + Lower (H:K)
    ws_matrix.cell(row_offset, 12, f"=SUM(D{row_offset}:K{row_offset})").font = font_bold
    ws_matrix.cell(row_offset, 13, f"=L{row_offset}").font = font_bold
    
    # Bath sub vanities
    ws_matrix.cell(row_offset, 14, row[11]).font = font_regular
    ws_matrix.cell(row_offset, 15, row[12]).font = font_regular
    ws_matrix.cell(row_offset, 16, row[13]).font = font_regular
    
    # Bath Sub = Sum of N:P
    ws_matrix.cell(row_offset, 17, f"=SUM(N{row_offset}:P{row_offset})").font = font_bold
    
    # Unit total = Kitchen Total + Bath Sub
    ws_matrix.cell(row_offset, 18, f"=M{row_offset}+Q{row_offset}").font = font_bold
    
    # Style and border
    for col_idx in range(1, 19):
        cell = ws_matrix.cell(row_offset, col_idx)
        cell.border = thin_border
        if col_idx in [1, 3]:
            cell.alignment = align_left
        else:
            cell.alignment = align_right

# Totals row
tot_row = 20
ws_matrix.cell(tot_row, 1, "PROJECT TOTALS").font = font_bold
ws_matrix.cell(tot_row, 2, "=SUM(B6:B19)").font = font_bold
ws_matrix.cell(tot_row, 3, "").font = font_regular

for col_idx in range(4, 19):
    col_letter = get_column_letter(col_idx)
    ws_matrix.cell(tot_row, col_idx, f"=SUM({col_letter}6:{col_letter}19)").font = font_bold

for col_idx in range(1, 19):
    cell = ws_matrix.cell(tot_row, col_idx)
    cell.border = double_bottom
    cell.fill = fill_light_gray
    if col_idx in [1, 3]:
        cell.alignment = align_left
    else:
        cell.alignment = align_right

# -------------------------------------------------------------
# SHOP DRAWING SCHEDULES PER UNIT TYPE
# -------------------------------------------------------------
# Standard cabinet libraries prices (calculated in USD with 1.7820 exchange rate)
usd_catalog = {
    # Upper
    "W3630": round(72.9 * 1.782, 2),
    "W3030": round(66.9 * 1.782, 2),
    "W-REF": round(45.1 * 1.782, 2), # Over fridge wall cab
    "W-MWO": round(44.1 * 1.782, 2), # Over microwave wall cab
    # Base/Lower
    "B36": round(73.0 * 1.782, 2),
    "B-30-DW": round(66.7 * 1.782, 2), # Base DW panel
    "B-SINK": round(90.0 * 1.782, 2),  # Sink base 36"
    "B24": round(54.1 * 1.782, 2),
    "B-RO": round(129.0 * 1.782, 2),   # Base roll out drawers
    # Tall
    "T-PANT": round(121.3 * 1.782, 2), # Pantry tall cab
    # Bathroom USD Standard Catalog
    "VAN-60": 379.0,
    "VAN-36": 249.0,
    "VAN-30": 219.0,
    "MED-60": 179.0,
    "MED-36": 119.0,
    "MED-30": 99.0,
    "LIN-18": 199.0,
}

unit_types = [
    "A-1", "A-1a ACC", "B-1", "B-1a ACC", "C-1", "C-2", "C-2a ACC", "C-3",
    "D-1N", "D-1", "D-1a ACC", "D-1A ACC", "ST-1a ACC", "ST-1"
]

unit_cabinet_lists = {
    "ST-1": [
        ("Upper", "W3030", "Wall Cabinet — 30\"W x 30\"H x 12\"D", 30, 30, 12, 1, "Elev.1", "Standard upper"),
        ("Upper", "W-REF", "Wall Cabinet — 30\"W x 12\"H x 12\"D (Fridge)", 30, 12, 12, 1, "Elev.1", "Over fridge"),
        ("Upper", "W-MWO", "Wall Cabinet — 30\"W x 18\"H x 12\"D (Microwave)", 30, 18, 12, 1, "Elev.1", "Over microwave"),
        ("Lower", "B-30-DW", "Dishwasher Cabinet space", 30, 34.5, 24, 1, "Elev.2", "DW adjacent"),
        ("Lower", "B-SINK", "Sink Base Cabinet — 36\"W", 36, 34.5, 24, 1, "Elev.2", "Sink base"),
        ("Lower", "B24", "Base Cabinet — 24\"W", 24, 34.5, 24, 1, "Elev.2", "Standard base"),
        ("Vanity", "VAN-30", "Bathroom Vanity — 30\"W", 30, 34.5, 21, 1, "Floor Plan", "Single sink"),
        ("Bath", "MED-30", "Medicine Cabinet — 30\"W", 30, 30, 4, 1, "Floor Plan", "Mirror cabinet")
    ],
    "ST-1a ACC": [
        ("Upper", "W3030", "Wall Cabinet — 30\"W x 30\"H x 12\"D", 30, 30, 12, 1, "Elev.1", "Standard upper"),
        ("Upper", "W-REF", "Wall Cabinet — 30\"W x 12\"H x 12\"D (Fridge)", 30, 12, 12, 1, "Elev.1", "Over fridge"),
        ("Upper", "W-MWO", "Wall Cabinet — 30\"W x 18\"H x 12\"D (Microwave)", 30, 18, 12, 1, "Elev.1", "Over microwave"),
        ("Lower", "B-30-DW", "Dishwasher Cabinet space (Accessible)", 30, 34.5, 24, 1, "Elev.2", "ADA DW adjacent"),
        ("Lower", "B-SINK", "Sink Base Cabinet — 36\"W (Accessible)", 36, 34.5, 24, 1, "Elev.2", "ADA Sink base"),
        ("Lower", "B24", "Base Cabinet — 24\"W (Accessible)", 24, 34.5, 24, 1, "Elev.2", "ADA Standard base"),
        ("Vanity", "VAN-30", "Bathroom Vanity — 30\"W (Accessible)", 30, 34.5, 21, 1, "Floor Plan", "ADA Single sink"),
        ("Bath", "MED-30", "Medicine Cabinet — 30\"W", 30, 30, 4, 1, "Floor Plan", "Mirror cabinet")
    ],
    "A-1": [
        ("Upper", "W3630", "Wall Cabinet — 36\"W x 30\"H x 12\"D", 36, 30, 12, 2, "Elev.1", "Standard upper"),
        ("Upper", "W-REF", "Wall Cabinet — 30\"W x 12\"H x 12\"D (Fridge)", 30, 12, 12, 1, "Elev.1", "Over fridge"),
        ("Upper", "W-MWO", "Wall Cabinet — 30\"W x 18\"H x 12\"D (Microwave)", 30, 18, 12, 1, "Elev.1", "Over microwave"),
        ("Upper", "W3030", "Wall Cabinet — 30\"W x 30\"H x 12\"D", 30, 30, 12, 1, "Elev.1", "End upper"),
        ("Lower", "B36", "Base Cabinet — 36\"W", 36, 34.5, 24, 2, "Elev.2", "Standard base"),
        ("Lower", "B-30-DW", "Dishwasher Cabinet space", 30, 34.5, 24, 1, "Elev.2", "DW space"),
        ("Lower", "B-SINK", "Sink Base Cabinet — 36\"W", 36, 34.5, 24, 1, "Elev.2", "Sink base"),
        ("Lower", "B24", "Base Cabinet — 24\"W", 24, 34.5, 24, 1, "Elev.2", "Standard base"),
        ("Tall", "T-PANT", "Tall Pantry Cabinet — 15\"W", 15, 84, 24, 1, "Elev.2", "Tall pantry"),
        ("Vanity", "VAN-36", "Bathroom Vanity — 36\"W", 36, 34.5, 21, 1, "Floor Plan", "Single sink"),
        ("Bath", "MED-36", "Medicine Cabinet — 36\"W", 36, 30, 4, 1, "Floor Plan", "Mirror cabinet"),
        ("Bath", "LIN18", "Linen Cabinet — 18\"W", 18, 84, 18, 1, "Elev.3", "Linen storage")
    ],
    "A-1a ACC": [
        ("Upper", "W3630", "Wall Cabinet — 36\"W x 30\"H x 12\"D", 36, 30, 12, 2, "Elev.1", "Standard upper"),
        ("Upper", "W-REF", "Wall Cabinet — 30\"W x 12\"H x 12\"D (Fridge)", 30, 12, 12, 1, "Elev.1", "Over fridge"),
        ("Upper", "W-MWO", "Wall Cabinet — 30\"W x 18\"H x 12\"D (Microwave)", 30, 18, 12, 1, "Elev.1", "Over microwave"),
        ("Upper", "W3030", "Wall Cabinet — 30\"W x 30\"H x 12\"D", 30, 30, 12, 1, "Elev.1", "End upper"),
        ("Lower", "B36", "Base Cabinet — 36\"W (Accessible)", 36, 34.5, 24, 2, "Elev.2", "ADA base"),
        ("Lower", "B-30-DW", "Dishwasher Cabinet space (Accessible)", 30, 34.5, 24, 1, "Elev.2", "ADA DW space"),
        ("Lower", "B-SINK", "Sink Base Cabinet — 36\"W (Accessible)", 36, 34.5, 24, 1, "Elev.2", "ADA Sink base"),
        ("Lower", "B24", "Base Cabinet — 24\"W (Accessible)", 24, 34.5, 24, 1, "Elev.2", "ADA base"),
        ("Tall", "T-PANT", "Tall Pantry Cabinet — 15\"W", 15, 84, 24, 1, "Elev.2", "Tall pantry"),
        ("Vanity", "VAN-36", "Bathroom Vanity — 36\"W (Accessible)", 36, 34.5, 21, 1, "Floor Plan", "ADA Single sink"),
        ("Bath", "MED-36", "Medicine Cabinet — 36\"W", 36, 30, 4, 1, "Floor Plan", "Mirror cabinet"),
        ("Bath", "LIN18", "Linen Cabinet — 18\"W", 18, 84, 18, 1, "Elev.3", "Linen storage")
    ],
    "B-1": [
        ("Upper", "W3630", "Wall Cabinet — 36\"W x 30\"H x 12\"D", 36, 30, 12, 2, "Elev.1", "Standard upper"),
        ("Upper", "W3030", "Wall Cabinet — 30\"W x 30\"H x 12\"D", 30, 30, 12, 2, "Elev.1", "Wall cabinets"),
        ("Upper", "W-REF", "Wall Cabinet — 30\"W x 12\"H x 12\"D (Fridge)", 30, 12, 12, 1, "Elev.1", "Over fridge"),
        ("Upper", "W-MWO", "Wall Cabinet — 30\"W x 18\"H x 12\"D (Microwave)", 30, 18, 12, 1, "Elev.1", "Over microwave"),
        ("Lower", "B36", "Base Cabinet — 36\"W", 36, 34.5, 24, 2, "Elev.2", "Standard base"),
        ("Lower", "B-30-DW", "Dishwasher Cabinet space", 30, 34.5, 24, 1, "Elev.2", "DW space"),
        ("Lower", "B-SINK", "Sink Base Cabinet — 36\"W", 36, 34.5, 24, 1, "Elev.2", "Sink base"),
        ("Lower", "B24", "Base Cabinet — 24\"W", 24, 34.5, 24, 2, "Elev.2", "Base cabinets"),
        ("Tall", "T-PANT", "Tall Pantry Cabinet — 15\"W", 15, 84, 24, 1, "Elev.2", "Tall pantry"),
        ("Vanity", "VAN-30", "Bathroom Vanity — 30\"W", 30, 34.5, 21, 2, "Floor Plan", "Single sink"),
        ("Bath", "MED-30", "Medicine Cabinet — 30\"W", 30, 30, 4, 2, "Floor Plan", "Mirror cabinet"),
        ("Bath", "LIN18", "Linen Cabinet — 18\"W", 18, 84, 18, 1, "Elev.3", "Linen storage")
    ],
    "B-1a ACC": [
        ("Upper", "W3630", "Wall Cabinet — 36\"W x 30\"H x 12\"D", 36, 30, 12, 2, "Elev.1", "Standard upper"),
        ("Upper", "W3030", "Wall Cabinet — 30\"W x 30\"H x 12\"D", 30, 30, 12, 2, "Elev.1", "Wall cabinets"),
        ("Upper", "W-REF", "Wall Cabinet — 30\"W x 12\"H x 12\"D (Fridge)", 30, 12, 12, 1, "Elev.1", "Over fridge"),
        ("Upper", "W-MWO", "Wall Cabinet — 30\"W x 18\"H x 12\"D (Microwave)", 30, 18, 12, 1, "Elev.1", "Over microwave"),
        ("Lower", "B36", "Base Cabinet — 36\"W (Accessible)", 36, 34.5, 24, 2, "Elev.2", "ADA base"),
        ("Lower", "B-30-DW", "Dishwasher Cabinet space (Accessible)", 30, 34.5, 24, 1, "Elev.2", "ADA DW space"),
        ("Lower", "B-SINK", "Sink Base Cabinet — 36\"W (Accessible)", 36, 34.5, 24, 1, "Elev.2", "ADA Sink base"),
        ("Lower", "B24", "Base Cabinet — 24\"W (Accessible)", 24, 34.5, 24, 2, "Elev.2", "ADA base"),
        ("Tall", "T-PANT", "Tall Pantry Cabinet — 15\"W", 15, 84, 24, 1, "Elev.2", "Tall pantry"),
        ("Vanity", "VAN-30", "Bathroom Vanity — 30\"W (Accessible)", 30, 34.5, 21, 2, "Floor Plan", "ADA Single sink"),
        ("Bath", "MED-30", "Medicine Cabinet — 30\"W", 30, 30, 4, 2, "Floor Plan", "Mirror cabinet"),
        ("Bath", "LIN18", "Linen Cabinet — 18\"W", 18, 84, 18, 1, "Elev.3", "Linen storage")
    ]
}

# Copy configurations for C-1 to C-3 and D-1
# C-1 to C-3 are identical
for c_type in ["C-1", "C-2", "C-2a ACC", "C-3"]:
    is_acc = "ACC" in c_type or "a " in c_type
    suffix = " (Accessible)" if is_acc else ""
    unit_cabinet_lists[c_type] = [
        ("Upper", "W3630", "Wall Cabinet — 36\"W x 30\"H x 12\"D", 36, 30, 12, 3, "Elev.1", "Standard upper"),
        ("Upper", "W-REF", "Wall Cabinet — 30\"W x 12\"H x 12\"D (Fridge)", 30, 12, 12, 1, "Elev.1", "Over fridge"),
        ("Upper", "W-MWO", "Wall Cabinet — 30\"W x 18\"H x 12\"D (Microwave)", 30, 18, 12, 1, "Elev.1", "Over microwave"),
        ("Upper", "W3030", "Wall Cabinet — 30\"W x 30\"H x 12\"D", 30, 30, 12, 1, "Elev.1", "Wall cabinet"),
        ("Lower", "B36", f"Base Cabinet — 36\"W{suffix}", 36, 34.5, 24, 3, "Elev.2", "Base run"),
        ("Lower", "B-30-DW", f"Dishwasher Cabinet space{suffix}", 30, 34.5, 24, 1, "Elev.2", "DW space"),
        ("Lower", "B-SINK", f"Sink Base Cabinet — 36\"W{suffix}", 36, 34.5, 24, 1, "Elev.2", "Sink base"),
        ("Lower", "B24", f"Base Cabinet — 24\"W{suffix}", 24, 34.5, 24, 1, "Elev.2", "Base cabinet"),
        ("Tall", "T-PANT", "Tall Pantry Cabinet — 15\"W", 15, 84, 24, 1, "Elev.2", "Tall pantry"),
        ("Vanity", "VAN-36", f"Bathroom Vanity — 36\"W{suffix}", 36, 34.5, 21, 1, "Floor Plan", "Vanity run"),
        ("Vanity", "VAN-30", f"Bathroom Vanity — 30\"W{suffix}", 30, 34.5, 21, 1, "Floor Plan", "Vanity run"),
        ("Bath", "MED-36", "Medicine Cabinet — 36\"W", 36, 30, 4, 1, "Floor Plan", "Mirror cabinet"),
        ("Bath", "MED-30", "Medicine Cabinet — 30\"W", 30, 30, 4, 1, "Floor Plan", "Mirror cabinet"),
        ("Bath", "LIN18", "Linen Cabinet — 18\"W", 18, 84, 18, 1, "Elev.3", "Linen storage")
    ]

# D-1 types (D-1, D-1N, D-1a ACC, D-1A ACC)
for d_type in ["D-1", "D-1N", "D-1a ACC", "D-1A ACC"]:
    is_acc = "ACC" in d_type or "a " in d_type
    suffix = " (Accessible)" if is_acc else ""
    unit_cabinet_lists[d_type] = [
        ("Upper", "W3630", "Wall Cabinet — 36\"W x 30\"H x 12\"D", 36, 30, 12, 3, "Elev.1", "Standard upper"),
        ("Upper", "W3030", "Wall Cabinet — 30\"W x 30\"H x 12\"D", 30, 30, 12, 2, "Elev.1", "Standard upper"),
        ("Upper", "W-REF", "Wall Cabinet — 30\"W x 12\"H x 12\"D (Fridge)", 30, 12, 12, 1, "Elev.1", "Over fridge"),
        ("Upper", "W-MWO", "Wall Cabinet — 30\"W x 18\"H x 12\"D (Microwave)", 30, 18, 12, 1, "Elev.1", "Over microwave"),
        ("Lower", "B36", f"Base Cabinet — 36\"W{suffix}", 36, 34.5, 24, 3, "Elev.2", "Base run"),
        ("Lower", "B-30-DW", f"Dishwasher Cabinet space{suffix}", 30, 34.5, 24, 1, "Elev.2", "DW space"),
        ("Lower", "B-SINK", f"Sink Base Cabinet — 36\"W{suffix}", 36, 34.5, 24, 1, "Elev.2", "Sink base"),
        ("Lower", "B24", f"Base Cabinet — 24\"W{suffix}", 24, 34.5, 24, 1, "Elev.2", "Base cabinet"),
        ("Lower", "B-RO", f"Base Cabinet with Roll-out Drawers{suffix}", 36, 34.5, 24, 1, "Elev.2", "Roll-out drawers"),
        ("Tall", "T-PANT", "Tall Pantry Cabinet — 15\"W", 15, 84, 24, 1, "Elev.2", "Tall pantry"),
        ("Vanity", "VAN-60", f"Bathroom Vanity — 60\"W{suffix}", 60, 34.5, 21, 1, "Floor Plan", "Double sink vanity"),
        ("Vanity", "VAN-36", f"Bathroom Vanity — 36\"W{suffix}", 36, 34.5, 21, 1, "Floor Plan", "Single sink vanity"),
        ("Bath", "MED-60", "Medicine Cabinet — 60\"W", 60, 30, 4, 1, "Floor Plan", "Double mirror"),
        ("Bath", "MED-36", "Medicine Cabinet — 36\"W", 36, 30, 4, 1, "Floor Plan", "Single mirror"),
        ("Bath", "LIN18", "Linen Cabinet — 18\"W", 18, 84, 18, 1, "Elev.3", "Linen storage")
    ]

# Populate worksheets
unit_material_costs = {}

for ut in unit_types:
    ws_shop = wb.create_sheet(f"Shop Drawing - {ut}")
    ws_shop.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_shop.cell(1, 1, "CABINET SHOP DRAWING — UNIT SCHEDULE").font = font_title
    ws_shop.cell(2, 1, "Atlantic Pacific Communities  |  Corwil Architects").font = font_small
    
    ws_shop.cell(4, 1, "PROJECT:").font = font_bold
    ws_shop.cell(4, 2, "Heritage Village").font = font_regular
    ws_shop.cell(4, 5, "UNIT TYPE:").font = font_bold
    ws_shop.cell(4, 6, ut).font = font_bold
    ws_shop.cell(4, 9, "DATE:").font = font_bold
    ws_shop.cell(4, 10, "=TODAY()").font = font_regular
    
    ws_shop.cell(5, 1, "ADDRESS:").font = font_bold
    ws_shop.cell(5, 2, "26905 SW 142nd Ave, Homestead FL").font = font_regular
    ws_shop.cell(5, 5, "DESCRIPTION:").font = font_bold
    
    # Determine unit description
    desc_str = "1-Bedroom / 1-Bath — FHA Type B"
    if "ACC" in ut or "a " in ut:
        if "ST" in ut:
            desc_str = "Studio Apartment — Fully Accessible ADA"
        elif "A-1" in ut:
            desc_str = "1-Bedroom / 1-Bath — Fully Accessible ADA"
        elif "B-1" in ut:
            desc_str = "2-Bedroom / 2-Bath — Fully Accessible ADA"
        elif "C-2" in ut:
            desc_str = "3-Bedroom / 2-Bath — Fully Accessible ADA"
        elif "D-1" in ut:
            desc_str = "4-Bedroom / 2-Bath — Fully Accessible ADA"
    else:
        if "ST" in ut:
            desc_str = "Studio Apartment — FHA Type B"
        elif "B-1" in ut:
            desc_str = "2-Bedroom / 2-Bath — FHA Type B"
        elif "C-1" in ut or "C-2" in ut or "C-3" in ut:
            desc_str = "3-Bedroom / 2-Bath — FHA Type B"
        elif "D-1" in ut or "D-1N" in ut:
            desc_str = "4-Bedroom / 2-Bath — FHA Type B"
            
    ws_shop.cell(5, 6, desc_str).font = font_regular
    ws_shop.cell(5, 9, "PREPARED BY:").font = font_bold
    ws_shop.cell(5, 10, "AI Estimation System").font = font_regular
    
    ws_shop.cell(6, 1, "DRAWING REF:").font = font_bold
    ws_shop.cell(6, 2, "Heritage Unit Plans Elevation Files").font = font_regular
    ws_shop.cell(6, 5, "FINISH TIER:").font = font_bold
    ws_shop.cell(6, 6, "Standard 1").font = font_regular
    
    ws_shop.cell(8, 1, "SECTION 1 — KITCHEN CABINETS").font = font_section
    
    # Headers
    headers = ["Item", "Cabinet Code", "Description", "Type", "W (in)", "H (in)", "D (in)", "Qty", "Finish Tier", "Elevation Ref.", "Location Note"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_shop.cell(10, col_idx, h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center

    cab_list = unit_cabinet_lists[ut]
    row_idx = 11
    kitchen_qty = 0
    bath_qty = 0
    
    kitchen_cost = 0
    bath_cost = 0
    
    item_num = 1
    
    # Write Kitchen Cabinets
    for cab in cab_list:
        if cab[0] in ["Upper", "Lower", "Tall"]:
            ws_shop.cell(row_idx, 1, item_num).font = font_regular
            ws_shop.cell(row_idx, 2, cab[1]).font = font_bold
            ws_shop.cell(row_idx, 3, cab[2]).font = font_regular
            ws_shop.cell(row_idx, 4, cab[0]).font = font_regular
            ws_shop.cell(row_idx, 5, cab[3]).font = font_regular
            ws_shop.cell(row_idx, 6, cab[4]).font = font_regular
            ws_shop.cell(row_idx, 7, cab[5]).font = font_regular
            ws_shop.cell(row_idx, 8, cab[6]).font = font_regular
            ws_shop.cell(row_idx, 9, "Std 1").font = font_regular
            ws_shop.cell(row_idx, 10, cab[7]).font = font_regular
            ws_shop.cell(row_idx, 11, cab[8]).font = font_regular
            
            for col in range(1, 12):
                cell = ws_shop.cell(row_idx, col)
                cell.border = thin_border
                if col in [1, 2, 3, 4, 9, 10, 11]:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_right
                    
            kitchen_qty += cab[6]
            kitchen_cost += usd_catalog.get(cab[1], 100.0) * cab[6]
            item_num += 1
            row_idx += 1
            
    # Kitchen Subtotals row
    ws_shop.cell(row_idx, 1, "KITCHEN SUBTOTALS").font = font_bold
    ws_shop.cell(row_idx, 8, kitchen_qty).font = font_bold
    ws_shop.cell(row_idx, 1, "KITCHEN SUBTOTALS").alignment = align_left
    ws_shop.cell(row_idx, 8).alignment = align_right
    for col in range(1, 12):
        cell = ws_shop.cell(row_idx, col)
        cell.border = double_bottom
        cell.fill = fill_light_gray
    row_idx += 2
    
    # Section 2: Bathroom Vanities
    ws_shop.cell(row_idx, 1, "SECTION 2 — BATHROOM VANITY CABINETS").font = font_section
    row_idx += 2
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws_shop.cell(row_idx, col_idx, h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
    row_idx += 1
    
    item_num_bath = 1
    for cab in cab_list:
        if cab[0] in ["Vanity", "Bath"]:
            ws_shop.cell(row_idx, 1, item_num_bath).font = font_regular
            ws_shop.cell(row_idx, 2, cab[1]).font = font_bold
            ws_shop.cell(row_idx, 3, cab[2]).font = font_regular
            ws_shop.cell(row_idx, 4, cab[0]).font = font_regular
            ws_shop.cell(row_idx, 5, cab[3]).font = font_regular
            ws_shop.cell(row_idx, 6, cab[4]).font = font_regular
            ws_shop.cell(row_idx, 7, cab[5]).font = font_regular
            ws_shop.cell(row_idx, 8, cab[6]).font = font_regular
            ws_shop.cell(row_idx, 9, "Std 1").font = font_regular
            ws_shop.cell(row_idx, 10, cab[7]).font = font_regular
            ws_shop.cell(row_idx, 11, cab[8]).font = font_regular
            
            for col in range(1, 12):
                cell = ws_shop.cell(row_idx, col)
                cell.border = thin_border
                if col in [1, 2, 3, 4, 9, 10, 11]:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_right
                    
            bath_qty += cab[6]
            bath_cost += usd_catalog.get(cab[1], 100.0) * cab[6]
            item_num_bath += 1
            row_idx += 1
            
    # Bath Subtotals row
    ws_shop.cell(row_idx, 1, "BATHROOM SUBTOTALS").font = font_bold
    ws_shop.cell(row_idx, 8, bath_qty).font = font_bold
    ws_shop.cell(row_idx, 1, "BATHROOM SUBTOTALS").alignment = align_left
    ws_shop.cell(row_idx, 8).alignment = align_right
    for col in range(1, 12):
        cell = ws_shop.cell(row_idx, col)
        cell.border = double_bottom
        cell.fill = fill_light_gray
    row_idx += 2
    
    # Grand Total row
    ws_shop.cell(row_idx, 1, f"UNIT {ut} GRAND TOTAL CABINET COUNT").font = font_bold
    ws_shop.cell(row_idx, 8, kitchen_qty + bath_qty).font = font_bold
    ws_shop.cell(row_idx, 1).alignment = align_left
    ws_shop.cell(row_idx, 8).alignment = align_right
    for col in range(1, 12):
        cell = ws_shop.cell(row_idx, col)
        cell.border = double_bottom
        cell.fill = fill_light_gray
        
    unit_material_costs[ut] = (kitchen_cost, bath_cost)
    
    # Auto column widths
    for col in ws_shop.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.value is not None and not val_str.startswith('='):
                max_len = max(max_len, len(val_str))
        ws_shop.column_dimensions[col_letter].width = max(max_len + 3, 10)

# -------------------------------------------------------------
# SHEET 4: Job Costing
# -------------------------------------------------------------
ws_costing.views.sheetView[0].showGridLines = True
ws_costing.cell(1, 1, "JOB COSTING — HERITAGE VILLAGE TOTAL ESTIMATE").font = font_title
ws_costing.cell(2, 1, "Atlantic Pacific Communities  |  GP Target: 35%").font = font_small

# Calculate total project kitchen & bath costs
total_kitchen_cost_all = 0
total_bath_cost_all = 0
total_kitchen_qty_all = 0
total_bath_qty_all = 0

for row_idx in range(6, 20):
    ut = ws_matrix.cell(row_idx, 1).value
    qty = ws_matrix.cell(row_idx, 2).value
    # Remove description details from name if there
    ut_clean = ut.split(' — ')[0].split(' \u2014 ')[0].strip()
    
    k_cost_unit, b_cost_unit = unit_material_costs[ut_clean]
    
    total_kitchen_cost_all += k_cost_unit * qty
    total_bath_cost_all += b_cost_unit * qty
    
    # Extract quantities
    k_qty_unit = ws_matrix.cell(row_idx, 13).value
    b_qty_unit = ws_matrix.cell(row_idx, 17).value
    
    # Simple hardcoded counts mapping for calculations (from matrix definition)
    k_qty_map = {
        "A-1": 8, "A-1a ACC": 8, "B-1": 11, "B-1a ACC": 11, "C-1": 11, "C-2": 11,
        "C-2a ACC": 11, "C-3": 11, "D-1N": 13, "D-1": 14, "D-1a ACC": 14, "D-1A ACC": 14,
        "ST-1a ACC": 6, "ST-1": 6
    }
    b_qty_map = {
        "A-1": 3, "A-1a ACC": 3, "B-1": 5, "B-1a ACC": 5, "C-1": 5, "C-2": 5,
        "C-2a ACC": 5, "C-3": 5, "D-1N": 5, "D-1": 5, "D-1a ACC": 5, "D-1A ACC": 5,
        "ST-1a ACC": 2, "ST-1": 2
    }
    
    total_kitchen_qty_all += k_qty_map[ut_clean] * qty
    total_bath_qty_all += b_qty_map[ut_clean] * qty

grand_material_cost = total_kitchen_cost_all + total_bath_cost_all
grand_cabinet_count = total_kitchen_qty_all + total_bath_qty_all

# Populate Job Costing Rows
ws_costing.cell(4, 1, "A. QUANTITIES").font = font_bold
ws_costing.cell(5, 2, "Kitchen Cabinets").font = font_regular
ws_costing.cell(5, 5, total_kitchen_qty_all).font = font_bold
ws_costing.cell(5, 6, "units").font = font_regular

ws_costing.cell(6, 2, "Bath Cabinets").font = font_regular
ws_costing.cell(6, 5, total_bath_qty_all).font = font_bold
ws_costing.cell(6, 6, "units").font = font_regular

ws_costing.cell(7, 2, "TOTAL CABINETS").font = font_bold
ws_costing.cell(7, 5, "=E5+E6").font = font_bold
ws_costing.cell(7, 6, "units").font = font_regular

ws_costing.cell(8, 2, "Estimated Containers").font = font_regular
ws_costing.cell(8, 5, "=ROUNDUP(E7/220,0)").font = font_bold
ws_costing.cell(8, 6, "containers").font = font_regular

ws_costing.cell(10, 1, "B. MATERIAL COST").font = font_bold
ws_costing.cell(11, 2, "Kitchen Cabinets Cost").font = font_regular
ws_costing.cell(11, 5, total_kitchen_cost_all).font = font_regular
ws_costing.cell(11, 5).number_format = "$#,##0"

ws_costing.cell(12, 2, "Bath Cabinets Cost").font = font_regular
ws_costing.cell(12, 5, total_bath_cost_all).font = font_regular
ws_costing.cell(12, 5).number_format = "$#,##0"

ws_costing.cell(13, 2, "MATERIAL SUBTOTAL").font = font_bold
ws_costing.cell(13, 5, "=E11+E12").font = font_bold
ws_costing.cell(13, 5).number_format = "$#,##0"

# C. Cost Build-Up
ws_costing.cell(15, 1, "C. PROJECT COSTS").font = font_bold
cost_items = [
    ("1", "Material Cost (cabinets)", "Invoice / Price List", "", "=E13"),
    ("2", "Local Use Tax", "% of Material Cost", 0.075, "=E21*D22"),
    ("3", "Ocean Freight / Shipping", "Per Container", 4500, "=E8*D23"),
    ("4", "Inland Delivery", "Fixed per Project", 1200, "=D24"),
    ("5", "Installation", "Per Cabinet Installed", 85, "=E7*D25"),
    ("6", "Warehousing", "% of Material Cost", 0.02, "=E21*D26"),
    ("7", "Material Protection", "% of Material Cost", 0.005, "=E21*D27"),
    ("8", "Insurance", "% of Material Cost", 0.008, "=E21*D28"),
    ("9", "Miscellaneous Allowance", "Fixed", 500, "=D29"),
]

# Set headers for Cost buildup
ws_costing.cell(19, 1, "#").font = font_bold
ws_costing.cell(19, 2, "Cost Item").font = font_bold
ws_costing.cell(19, 3, "Basis / Driver").font = font_bold
ws_costing.cell(19, 4, "Rate").font = font_bold
ws_costing.cell(19, 5, "Calculated ($)").font = font_bold
ws_costing.cell(19, 6, "Notes").font = font_bold

for col in range(1, 7):
    ws_costing.cell(19, col).fill = fill_light_gray
    ws_costing.cell(19, col).border = thin_border

for idx, item in enumerate(cost_items, 21):
    ws_costing.cell(idx, 1, item[0]).font = font_regular
    ws_costing.cell(idx, 2, item[1]).font = font_regular
    ws_costing.cell(idx, 3, item[2]).font = font_regular
    cell_rate = ws_costing.cell(idx, 4, item[3])
    cell_rate.font = font_regular
    if isinstance(item[3], float) and item[3] < 1.0:
        cell_rate.number_format = "0.0%"
    elif isinstance(item[3], int) or isinstance(item[3], float):
        cell_rate.number_format = "$#,##0"
        
    cell_calc = ws_costing.cell(idx, 5, item[4])
    cell_calc.font = font_bold
    cell_calc.number_format = "$#,##0.00"
    
    for col in range(1, 7):
        ws_costing.cell(idx, col).border = thin_border

# Pre-margin subtotal
ws_costing.cell(31, 1, "PRE-MARGIN SUBTOTAL (Rows 1-9)").font = font_bold
ws_costing.cell(31, 5, "=SUM(E21:E29)").font = font_bold
ws_costing.cell(31, 5).number_format = "$#,##0.00"
for col in range(1, 7):
    ws_costing.cell(31, col).fill = fill_light_gray
    ws_costing.cell(31, col).border = thin_border

# Margin and Overhead
ws_costing.cell(33, 1, "D. MARGIN & OVERHEAD").font = font_bold
ws_costing.cell(34, 1, "10").font = font_regular
ws_costing.cell(34, 2, "Commission").font = font_regular
ws_costing.cell(34, 3, "% of Selling Price").font = font_regular
ws_costing.cell(34, 4, 0.05).font = font_regular
ws_costing.cell(34, 4).number_format = "0.0%"
ws_costing.cell(34, 5, "=E46*D34").font = font_bold
ws_costing.cell(34, 5).number_format = "$#,##0.00"

ws_costing.cell(35, 1, "11").font = font_regular
ws_costing.cell(35, 2, "Bond").font = font_regular
ws_costing.cell(35, 3, "% of Selling Price").font = font_regular
ws_costing.cell(35, 4, 0.015).font = font_regular
ws_costing.cell(35, 4).number_format = "0.0%"
ws_costing.cell(35, 5, "=E46*D35").font = font_bold
ws_costing.cell(35, 5).number_format = "$#,##0.00"

ws_costing.cell(36, 1, "12").font = font_regular
ws_costing.cell(36, 2, "Gross Profit").font = font_regular
ws_costing.cell(36, 3, "% of Selling Price").font = font_regular
ws_costing.cell(36, 4, 0.35).font = font_regular
ws_costing.cell(36, 4).number_format = "0.0%"
ws_costing.cell(36, 5, "=E46*D36").font = font_bold
ws_costing.cell(36, 5).number_format = "$#,##0.00"

for r in [34, 35, 36]:
    for c in range(1, 7):
        ws_costing.cell(r, c).border = thin_border

# Selling price calculation formula
ws_costing.cell(39, 1, "E. SELLING PRICE CALCULATION").font = font_bold
ws_costing.cell(40, 1, "Pre-Margin Cost:").font = font_bold
ws_costing.cell(40, 2, "=E31").font = font_regular
ws_costing.cell(40, 2).number_format = "$#,##0.00"

ws_costing.cell(41, 1, "GP Target %:").font = font_bold
ws_costing.cell(41, 2, "=D36").font = font_regular
ws_costing.cell(41, 2).number_format = "0.0%"

ws_costing.cell(42, 1, "Commission %:").font = font_bold
ws_costing.cell(42, 2, "=D34").font = font_regular
ws_costing.cell(42, 2).number_format = "0.0%"

ws_costing.cell(43, 1, "Bond %:").font = font_bold
ws_costing.cell(43, 2, "=D35").font = font_regular
ws_costing.cell(43, 2).number_format = "0.0%"

# SELLING PRICE (TOTAL PROJECT)
ws_costing.cell(46, 1, "SELLING PRICE (TOTAL PROJECT):").font = font_bold
ws_costing.cell(46, 5, "=IF(1-D36-D34-D35<=0,\"ERROR\",E31/(1-D36-D34-D35))").font = font_title
ws_costing.cell(46, 5).fill = fill_green_calc
ws_costing.cell(46, 5).number_format = "$#,##0.00"

# Verification
ws_costing.cell(48, 1, "VERIFICATION").font = font_bold
ws_costing.cell(49, 1, "Gross Profit $:").font = font_bold
ws_costing.cell(49, 5, "=E46*D36").font = font_bold
ws_costing.cell(49, 5).number_format = "$#,##0.00"

ws_costing.cell(50, 1, "Gross Profit %:").font = font_bold
ws_costing.cell(50, 5, "=IF(E46>0,E49/E46,0)").font = font_bold
ws_costing.cell(50, 5).number_format = "0.0%"

ws_costing.cell(51, 1, "Total Cost $:").font = font_bold
ws_costing.cell(51, 5, "=E46-E49").font = font_bold
ws_costing.cell(51, 5).number_format = "$#,##0.00"

ws_costing.cell(52, 1, "Cost per Cabinet:").font = font_bold
ws_costing.cell(52, 5, "=IF(E7>0,E51/E7,0)").font = font_bold
ws_costing.cell(52, 5).number_format = "$#,##0.00"

ws_costing.cell(53, 1, "Sell per Cabinet:").font = font_bold
ws_costing.cell(53, 5, "=IF(E7>0,E46/E7,0)").font = font_bold
ws_costing.cell(53, 5).number_format = "$#,##0.00"

for col in ws_costing.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        val_str = str(cell.value or '')
        if cell.value is not None and not val_str.startswith('='):
            max_len = max(max_len, len(val_str))
    ws_costing.column_dimensions[col_letter].width = max(max_len + 3, 15)

# Save the workbook
wb.save(OUTPUT_PATH)
print(f"Workbook compiled successfully and saved to: {OUTPUT_PATH}")
