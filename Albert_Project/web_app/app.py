"""
===========================================================================
  web_app/app.py  —  FastAPI Backend (Full Production)
===========================================================================
  Endpoints:
    GET  /api/projects                    list all projects
    POST /api/projects/create             create new project from wizard
    POST /api/projects/{id}/upload-pdf    upload architectural PDF for a unit
    POST /api/projects/{id}/upload-price-list  upload price list Excel
    POST /api/projects/{id}/save-config   update financial params
    GET  /api/projects/{id}/run           SSE pipeline stream
    GET  /api/projects/{id}/results       aggregated results + costing
    GET  /api/projects/{id}/status        quick status check
    DELETE /api/projects/{id}             delete project + outputs
    GET  /api/projects/{id}/download/pdf  serve shop drawing PDF
    GET  /api/projects/{id}/download/excel serve costing Excel
    GET  /api/projects/{id}/crops/{file}  serve elevation crop PNG
    GET  /api/projects/{id}/regions/{unit} detected region coords
    GET  /api/projects/{id}/pdf-pages/{unit} PDF thumbnail PNG
    GET  /                                serve static dashboard
===========================================================================
"""
from __future__ import annotations

import json
import re
import shutil
import sys
import os
import asyncio
import tempfile
from pathlib import Path
from typing import Any, Optional

import fitz  # PyMuPDF
from fastapi import FastAPI, HTTPException, Request, UploadFile, File, Form
from fastapi.responses import (
    FileResponse, StreamingResponse, Response, JSONResponse
)
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware

# ── Resolve project root ──────────────────────────────────────────────────────
_HERE        = Path(__file__).resolve().parent
PROJECT_ROOT = _HERE.parent

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from core.config import get_output_dir, OUTPUTS_ROOT, OPENROUTER_API_KEY  # noqa: E402
from core.ai_vision_classifier import PRIMARY_MODEL, FALLBACK_MODEL, DEMO_MODE  # noqa: E402

app = FastAPI(
    title       = "ItalianKB Cabinet Estimator",
    description = "AI-powered cabinet shop drawing and job costing automation",
    version     = "2.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

PROJECTS_ROOT = PROJECT_ROOT / "projects"
PROJECTS_ROOT.mkdir(exist_ok=True)
MAX_PDF_UPLOAD_BYTES = 100 * 1024 * 1024
MAX_EXCEL_UPLOAD_BYTES = 25 * 1024 * 1024

# ─────────────────────────────────────────────────────────────────────────────
# INTERNAL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _slugify(text: str) -> str:
    """Convert project name/id to a safe folder name."""
    return re.sub(r"[^a-zA-Z0-9_-]", "_", text.strip().lower())[:40]


def _validate_project_id(project_id: str) -> str:
    """Validate a project id before it is used for folders or output paths."""
    project_id = project_id.strip()
    if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9_.-]{1,40}", project_id):
        raise HTTPException(
            400,
            "project_id must be 2-41 characters and contain only letters, "
            "numbers, dot, underscore, or hyphen.",
        )
    if ".." in project_id:
        raise HTTPException(400, "project_id cannot contain '..'.")
    return project_id


def _safe_upload_name(name: str, fallback: str) -> str:
    """Return a safe basename for a user-supplied upload filename."""
    stem = Path(name or fallback).stem
    suffix = Path(name or fallback).suffix.lower()
    safe_stem = _slugify(stem) or fallback
    return f"{safe_stem}{suffix}"


def _ensure_inside(child: Path, parent: Path) -> Path:
    """Resolve a path and ensure it remains under the expected parent."""
    resolved = child.resolve()
    parent = parent.resolve()
    if parent not in resolved.parents and resolved != parent:
        raise HTTPException(400, "Resolved upload path escaped project folder.")
    return resolved


def _find_project(project_id: str) -> Optional[tuple[dict, Path, Path]]:
    """
    Search all project folders for one matching project_id.
    Returns (config_dict, config_path, project_folder) or None.
    """
    if not PROJECTS_ROOT.exists():
        return None
    for folder in PROJECTS_ROOT.iterdir():
        if not folder.is_dir():
            continue
        cfg_path = folder / "project_config.json"
        if not cfg_path.exists():
            continue
        try:
            cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
            if cfg.get("project_id") == project_id:
                return cfg, cfg_path, folder
        except Exception:
            continue
    return None


def _sse(obj: Any) -> str:
    return f"data: {json.dumps(obj, ensure_ascii=False)}\n\n"


def _default_price_list_path() -> Optional[str]:
    """Return relative path to default price list if one exists."""
    # Look for any Excel price list in known locations
    candidates = [
        PROJECT_ROOT / "Casa familia" / "02_Price_List" / "MS PRICE LIST LEVEL 1 -90CM.xlsx",
        PROJECT_ROOT / "Heritage" / "02_Price_List" / "MS PRICE LIST LEVEL 1 -90CM.xlsx",
    ]
    for c in candidates:
        if c.exists():
            return str(c.relative_to(PROJECT_ROOT))
    return None


@app.get("/api/health")
async def health():
    """Production readiness snapshot without exposing any secrets."""
    default_price_list = _default_price_list_path()
    project_count = 0
    if PROJECTS_ROOT.exists():
        project_count = sum(
            1 for f in PROJECTS_ROOT.iterdir()
            if f.is_dir() and (f / "project_config.json").exists()
        )
    return {
        "status": "ok",
        "version": app.version,
        "ai": {
            "provider": "OpenRouter",
            "primary_model": PRIMARY_MODEL,
            "fallback_model": FALLBACK_MODEL,
            "openrouter_key_configured": bool(OPENROUTER_API_KEY),
            "demo_mode_default": DEMO_MODE,
        },
        "storage": {
            "projects_root": str(PROJECTS_ROOT),
            "outputs_root": str(OUTPUTS_ROOT),
            "project_count": project_count,
            "default_price_list_available": bool(default_price_list),
            "default_price_list_path": default_price_list,
        },
        "limits": {
            "max_pdf_upload_mb": MAX_PDF_UPLOAD_BYTES // (1024 * 1024),
            "max_excel_upload_mb": MAX_EXCEL_UPLOAD_BYTES // (1024 * 1024),
        },
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/projects — list all
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/projects")
async def list_projects():
    results = []
    if not PROJECTS_ROOT.exists():
        return results
    for folder in sorted(PROJECTS_ROOT.iterdir()):
        if not folder.is_dir():
            continue
        cfg_path = folder / "project_config.json"
        if not cfg_path.exists():
            continue
        try:
            cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
            pid = cfg.get("project_id", "")
            out_dir = OUTPUTS_ROOT / pid
            results.append({
                "id":          pid,
                "name":        cfg.get("project_name", "Unnamed"),
                "folder_name": folder.name,
                "config":      cfg,
                "has_results": (out_dir / "json").exists() and
                               any((out_dir / "json").iterdir()),
                "has_pdf":     (out_dir / f"{pid}_Shop_Drawings.pdf").exists(),
                "has_excel":   (out_dir / f"{pid}_Cabinet_Estimation.xlsx").exists(),
            })
        except Exception as exc:
            print(f"[WARN] Cannot read {cfg_path}: {exc}")
    return results


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/projects/create — create new project from wizard
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/projects/create")
async def create_project(request: Request):
    """
    Create a new project from the wizard.
    Expects JSON body matching the project config schema.
    Automatically fills in price_list_path with the default if not provided.
    """
    try:
        body = await request.json()
    except Exception as exc:
        raise HTTPException(400, f"Invalid JSON: {exc}")

    project_id   = _validate_project_id(body.get("project_id", ""))
    project_name = body.get("project_name", "").strip()
    if not project_name:
        raise HTTPException(400, "project_name is required")

    # Check for duplicate
    if _find_project(project_id):
        raise HTTPException(409, f"Project ID '{project_id}' already exists.")

    # Create project folder
    folder_name = _slugify(project_id)
    project_folder = PROJECTS_ROOT / folder_name
    counter = 0
    while project_folder.exists():
        counter += 1
        project_folder = PROJECTS_ROOT / f"{folder_name}_{counter}"
    project_folder.mkdir(parents=True)
    (project_folder / "drawings").mkdir()

    # If no price list path provided, use the default ItalianKB list
    if not body.get("price_list_path"):
        default = _default_price_list_path()
        if default:
            body["price_list_path"] = default

    # Build full config
    config = {
        "project_id":         project_id,
        "project_name":       project_name,
        "date":               body.get("date", ""),
        "revision":           body.get("revision", "1.0"),
        "drawn_by":           body.get("drawn_by", ""),
        "checker":            body.get("checker", ""),
        "client_name":        body.get("client_name", ""),
        "architect":          body.get("architect", ""),
        "project_address":    body.get("project_address", ""),
        "cabinet_finish":     body.get("cabinet_finish", ""),
        "door_style":         body.get("door_style", ""),
        "finish_tier":        body.get("finish_tier", 1),

        "gp_target_pct":     float(body.get("gp_target_pct",  0.35)),
        "commission_pct":    float(body.get("commission_pct", 0.05)),
        "bond_pct":          float(body.get("bond_pct",       0.015)),
        "price_list_tier":   int(body.get("price_list_tier",  1)),
        "eur_usd_rate":      float(body.get("eur_usd_rate",   1.09)),

        "price_list_path":   body.get("price_list_path", ""),

        # Will be populated by upload-pdf endpoint
        "unit_plan_pdfs":    body.get("unit_plan_pdfs", {}),
        "unit_counts":       body.get("unit_counts", {}),
        "ada_units":         body.get("ada_units", []),

        "floor_plan_pdfs":   body.get("floor_plan_pdfs", []),

        "appliances_regular": body.get("appliances_regular", []),
        "appliances_ada":     body.get("appliances_ada", []),
        "general_notes":      body.get("general_notes", [
            "ALL CABINETS ARE 90CM DEPTH",
            "ALL HARDWARE CONCEALED",
            "SOFT CLOSE DOORS AND DRAWERS STANDARD",
        ]),

        "company": body.get("company", {
            "name":    "ITALIAN KITCHEN AND BATH",
            "tagline": "kitchen | bath | tile | closet",
            "address": "1777 NW 72TH AVE. MIAMI FL, 33126",
            "phone":   "T. 305.599.9000  F. 305.599.9870",
            "website": "www.italiankitchenandbath.com",
        }),
    }

    cfg_path = project_folder / "project_config.json"
    cfg_path.write_text(
        json.dumps(config, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    # Create output directory structure
    get_output_dir(project_id)

    return {
        "status":       "created",
        "project_id":   project_id,
        "folder":       project_folder.name,
        "config":       config,
    }


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/projects/{id}/upload-pdf — upload architectural PDF
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/projects/{project_id}/upload-pdf")
async def upload_pdf(
    project_id: str,
    unit_type:  str        = Form(...),
    is_ada:     bool       = Form(False),
    file:       UploadFile = File(...),
):
    """
    Upload one architectural PDF for a specific unit type.
    Saves to projects/{folder}/drawings/{unit_type}.pdf
    Updates project_config.json with the new unit_plan_pdfs entry.
    """
    result = _find_project(project_id)
    if result is None:
        raise HTTPException(404, f"Project '{project_id}' not found")
    cfg, cfg_path, project_folder = result

    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Only PDF files are accepted for unit drawings.")

    unit_type = unit_type.strip()
    if not unit_type:
        raise HTTPException(400, "unit_type is required.")

    # Save the PDF
    drawings_dir = project_folder / "drawings"
    drawings_dir.mkdir(exist_ok=True)

    safe_unit = _slugify(unit_type)
    if not safe_unit:
        raise HTTPException(400, "unit_type must contain letters or numbers.")
    dest = _ensure_inside(drawings_dir / f"{safe_unit}.pdf", drawings_dir)
    content = await file.read()
    if not content:
        raise HTTPException(400, "Uploaded PDF is empty.")
    if len(content) > MAX_PDF_UPLOAD_BYTES:
        raise HTTPException(413, "PDF upload exceeds 100 MB limit.")

    try:
        doc = fitz.open(stream=content, filetype="pdf")
        page_count = doc.page_count
        doc.close()
    except Exception as exc:
        raise HTTPException(400, f"Uploaded file is not a valid PDF: {exc}")

    dest.write_bytes(content)

    # Update config — store path RELATIVE to PROJECT_ROOT
    rel_path = str(dest.relative_to(PROJECT_ROOT))
    cfg.setdefault("unit_plan_pdfs", {})[unit_type] = rel_path

    # Track ADA status
    ada_units = set(cfg.get("ada_units", []))
    if is_ada:
        ada_units.add(unit_type)
    else:
        ada_units.discard(unit_type)
    cfg["ada_units"] = sorted(ada_units)

    cfg_path.write_text(
        json.dumps(cfg, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    return {
        "status":    "uploaded",
        "unit_type": unit_type,
        "path":      rel_path,
        "pages":     page_count,
        "size_kb":   round(len(content) / 1024, 1),
    }


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/projects/{id}/upload-price-list — upload Excel price list
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/projects/{project_id}/upload-price-list")
async def upload_price_list(
    project_id: str,
    file: UploadFile = File(...),
):
    result = _find_project(project_id)
    if result is None:
        raise HTTPException(404, f"Project '{project_id}' not found")
    cfg, cfg_path, project_folder = result

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx Excel files are accepted.")

    dest = _ensure_inside(project_folder / "price_list.xlsx", project_folder)
    content = await file.read()
    if not content:
        raise HTTPException(400, "Uploaded price list is empty.")
    if len(content) > MAX_EXCEL_UPLOAD_BYTES:
        raise HTTPException(413, "Price list upload exceeds 25 MB limit.")

    tmp_path = None
    try:
        from openpyxl import load_workbook
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        sheet_count = len(wb.sheetnames)
        wb.close()
        Path(tmp_path).unlink(missing_ok=True)
    except Exception as exc:
        if tmp_path:
            try:
                Path(tmp_path).unlink(missing_ok=True)
            except Exception:
                pass
        raise HTTPException(400, f"Uploaded file is not a valid .xlsx workbook: {exc}")

    dest.write_bytes(content)

    rel_path = str(dest.relative_to(PROJECT_ROOT))
    cfg["price_list_path"] = rel_path
    cfg_path.write_text(
        json.dumps(cfg, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    return {
        "status":  "uploaded",
        "path":    rel_path,
        "sheets":  sheet_count,
        "size_kb": round(len(content) / 1024, 1),
    }


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/projects/{id}/save-config — update project config
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/projects/{project_id}/save-config")
async def save_config(project_id: str, request: Request):
    result = _find_project(project_id)
    if result is None:
        raise HTTPException(404, f"Project '{project_id}' not found")
    _, cfg_path, _ = result

    try:
        new_config = await request.json()
    except Exception as exc:
        raise HTTPException(400, f"Invalid JSON body: {exc}")

    cfg_path.write_text(
        json.dumps(new_config, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    return {"status": "ok", "message": "Configuration saved."}


# ─────────────────────────────────────────────────────────────────────────────
# DELETE /api/projects/{id} — delete project + outputs
# ─────────────────────────────────────────────────────────────────────────────

@app.delete("/api/projects/{project_id}")
async def delete_project(project_id: str):
    result = _find_project(project_id)
    if result is None:
        raise HTTPException(404, f"Project '{project_id}' not found")
    _, _, project_folder = result

    # Remove project folder
    shutil.rmtree(project_folder, ignore_errors=True)

    # Remove outputs folder
    out_dir = OUTPUTS_ROOT / project_id
    if out_dir.exists():
        shutil.rmtree(out_dir, ignore_errors=True)

    return {"status": "deleted", "project_id": project_id}


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/projects/{id}/status — quick status check
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/projects/{project_id}/status")
async def project_status(project_id: str):
    result = _find_project(project_id)
    if result is None:
        raise HTTPException(404, f"Project '{project_id}' not found")
    cfg, _, _ = result
    out_dir = OUTPUTS_ROOT / project_id

    json_dir   = out_dir / "json"
    has_cache  = json_dir.exists() and any(json_dir.glob("cabinet_schedule_*.json"))
    has_pdf    = (out_dir / f"{project_id}_Shop_Drawings.pdf").exists()
    has_excel  = (out_dir / f"{project_id}_Cabinet_Estimation.xlsx").exists()

    unit_pdfs = cfg.get("unit_plan_pdfs", {})
    pdfs_ok = {
        u: (PROJECT_ROOT / p).exists()
        for u, p in unit_pdfs.items()
    }

    return {
        "project_id":   project_id,
        "project_name": cfg.get("project_name"),
        "units":        list(unit_pdfs.keys()),
        "pdf_files_ok": pdfs_ok,
        "ai_ready":     bool(OPENROUTER_API_KEY),
        "price_list_ok": bool(cfg.get("price_list_path")) and
                         (PROJECT_ROOT / cfg.get("price_list_path", "")).exists(),
        "has_cache":    has_cache,
        "has_pdf":      has_pdf,
        "has_excel":    has_excel,
        "ready_to_run": bool(unit_pdfs) and all(pdfs_ok.values()),
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/projects/{id}/run — run pipeline via SSE
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/projects/{project_id}/run")
async def run_pipeline(
    project_id: str,
    skip_ai:    bool = True,
    unit:       str  = None,
    demo:       bool = False,
):
    result = _find_project(project_id)
    if result is None:
        raise HTTPException(404, "Project not found")
    _, cfg_path, _ = result

    cmd = [sys.executable, str(PROJECT_ROOT / "pipeline.py"),
           "--project", str(cfg_path)]
    if skip_ai:
        cmd.append("--skip-ai")
    if unit:
        cmd += ["--unit", unit]
    if demo:
        cmd.append("--demo")

    async def generator():
        mode_label = "DEMO (low-token)" if demo else "FULL"
        launch_msg = f"[{mode_label}] Launching: " + " ".join(str(c) for c in cmd)
        yield _sse({"log": f"[SYSTEM] {launch_msg}"})

        try:
            env = {**os.environ, "PYTHONIOENCODING": "utf-8"}
            proc = await asyncio.create_subprocess_exec(
                *cmd,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.STDOUT,
                cwd=str(PROJECT_ROOT),
                env=env,
            )
        except Exception as exc:
            yield _sse({"log": f"[ERROR] Cannot start process: {exc}",
                        "done": True, "success": False})
            return

        while True:
            line = await proc.stdout.readline()
            if not line:
                break
            text = line.decode("utf-8", errors="replace").rstrip("\r\n")
            if text:
                yield _sse({"log": text})

        code = await proc.wait()
        yield _sse({
            "log":     f"[SYSTEM] Pipeline finished (exit {code}).",
            "done":    True,
            "success": code == 0,
        })

    return StreamingResponse(
        generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control":     "no-cache",
            "X-Accel-Buffering": "no",
        },
    )



# ─────────────────────────────────────────────────────────────────────────────
# GET /api/projects/{id}/results — aggregated results
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/projects/{project_id}/results")
async def get_results(project_id: str):
    out_dir = OUTPUTS_ROOT / project_id
    if not out_dir.exists():
        return JSONResponse(
            status_code=404,
            content={"detail": "No outputs yet — run the pipeline first."},
        )

    result = _find_project(project_id)
    if result is None:
        return JSONResponse(
            status_code=404,
            content={"detail": "Project config not found."},
        )
    config, _, _ = result

    # ── Load cached JSON schedules ──────────────────────────────────────────
    unit_schedules: dict = {}
    json_dir = out_dir / "json"
    if json_dir.exists():
        for f in json_dir.iterdir():
            if f.suffix != ".json" or not f.name.startswith("cabinet_schedule_"):
                continue
            raw_key = f.stem.replace("cabinet_schedule_", "")
            # Find matching key in config (handles underscore/hyphen normalization)
            unit_keys = list(config.get("unit_plan_pdfs", {}).keys())
            matched_key = raw_key
            for k in unit_keys:
                if k.replace(" ", "_").replace("-", "_") == raw_key:
                    matched_key = k
                    break
            try:
                unit_schedules[matched_key] = json.loads(
                    f.read_text(encoding="utf-8")
                )
            except Exception as exc:
                print(f"[WARN] Cannot parse {f.name}: {exc}")

    # ── Price matching ──────────────────────────────────────────────────────
    unit_totals = config.get("unit_counts", {})
    price_list  = PROJECT_ROOT / config.get("price_list_path", "")
    matcher     = None

    if price_list.exists():
        try:
            from core.price_matcher import PriceMatcher
            matcher = PriceMatcher(
                price_list,
                tier=config.get("price_list_tier", 1),
                eur_usd_rate=config.get("eur_usd_rate", 1.09),
            )
        except Exception as exc:
            print(f"[WARN] PriceMatcher failed: {exc}")

    from core.price_matcher import _generate_code, get_fallback_price_usd
    from core.ai_vision_classifier import CabinetItem

    total_material = 0.0
    total_cabs     = 0
    line_items     = []

    for unit_type, qty in unit_totals.items():
        sched = unit_schedules.get(unit_type)
        if not sched:
            continue
        for ev in sched.get("elevations", []):
            for cab_data in ev.get("cabinets", []):
                try:
                    cab = CabinetItem(**cab_data)
                except Exception:
                    continue
                if cab.cabinet_type == "appliance_space":
                    continue

                unit_price     = 0.0
                unit_price_eur = 0.0
                match_quality  = "NOT_FOUND"
                matched_desc   = cab.cabinet_type
                h_mm = cab.height_mm
                d_mm = cab.depth_mm
                cab_code = cab.code

                if matcher:
                    res = matcher.match(
                        cab.cabinet_type, cab.width_mm, quantity=cab.quantity
                    )
                    unit_price     = res.price_usd
                    unit_price_eur = res.price_eur
                    match_quality  = res.match_quality
                    if res.matched_entry:
                        matched_desc = res.matched_entry.description
                        h_mm         = res.matched_entry.height_mm
                        d_mm         = res.matched_entry.depth_mm
                else:
                    code           = _generate_code(cab.cabinet_type, cab.width_mm)
                    unit_price     = get_fallback_price_usd(code)
                    unit_price_eur = unit_price / config.get("eur_usd_rate", 1.09)
                    match_quality  = "FALLBACK"

                cab_count       = cab.quantity * qty
                total_material += unit_price * cab_count
                total_cabs     += cab_count

                line_items.append({
                    "unit_type":       unit_type,
                    "elevation":       ev.get("elevation_label"),
                    "item_num":        cab.item_num,
                    "type":            cab.cabinet_type,
                    "code":            cab_code,
                    "width_mm":        cab.width_mm,
                    "height_mm":       h_mm,
                    "depth_mm":        d_mm,
                    "confidence":      cab.confidence,
                    "quantity":        cab.quantity,
                    "unit_qty":        qty,
                    "total_qty":       cab_count,
                    "unit_price_usd":  unit_price,
                    "unit_price_eur":  unit_price_eur,
                    "total_price_usd": unit_price * cab_count,
                    "match_quality":   match_quality,
                    "matched_desc":    matched_desc,
                    "notes":           cab.notes,
                    "source":          cab.source,
                    "is_ada":          cab.is_ada,
                })

    validation_flags = []
    low_confidence_items = []
    invalid_dimension_items = []
    schedule_review_flags = []
    price_match_counts = {}

    for unit_type, sched in unit_schedules.items():
        for flag in sched.get("review_flags", []) or []:
            schedule_review_flags.append({"unit_type": unit_type, "flag": flag})
            validation_flags.append(f"{unit_type}: {flag}")
        for ev in sched.get("elevations", []) or []:
            elev_label = ev.get("elevation_label")
            for flag in ev.get("review_flags", []) or []:
                schedule_review_flags.append({
                    "unit_type": unit_type,
                    "elevation": elev_label,
                    "flag": flag,
                })
                validation_flags.append(f"{unit_type}/{elev_label}: {flag}")
            for cab in ev.get("cabinets", []) or []:
                conf = float(cab.get("confidence") or 0)
                if conf < 0.70:
                    low_confidence_items.append({
                        "unit_type": unit_type,
                        "elevation": elev_label,
                        "item_num": cab.get("item_num"),
                        "type": cab.get("cabinet_type"),
                        "confidence": conf,
                    })
                if any(float(cab.get(k) or 0) <= 0 for k in ("width_mm", "height_mm", "depth_mm")):
                    invalid_dimension_items.append({
                        "unit_type": unit_type,
                        "elevation": elev_label,
                        "item_num": cab.get("item_num"),
                        "type": cab.get("cabinet_type"),
                        "width_mm": cab.get("width_mm"),
                        "height_mm": cab.get("height_mm"),
                        "depth_mm": cab.get("depth_mm"),
                    })

    for item in line_items:
        quality = item.get("match_quality", "UNKNOWN")
        price_match_counts[quality] = price_match_counts.get(quality, 0) + 1
        if quality == "NOT_FOUND":
            validation_flags.append(
                f"{item.get('unit_type')}/{item.get('elevation')} item "
                f"{item.get('item_num')}: price match not found"
            )

    if invalid_dimension_items:
        validation_flags.append(
            f"{len(invalid_dimension_items)} cabinet item(s) have missing or zero dimensions"
        )
    if low_confidence_items:
        validation_flags.append(
            f"{len(low_confidence_items)} cabinet item(s) below 70% confidence"
        )

    validation = {
        "status": "PASS" if not validation_flags else "REVIEW",
        "flag_count": len(validation_flags),
        "flags": validation_flags[:50],
        "review_flags": schedule_review_flags,
        "low_confidence_items": low_confidence_items,
        "invalid_dimension_items": invalid_dimension_items,
        "price_match_counts": price_match_counts,
    }

    # ── Job costing ─────────────────────────────────────────────────────────
    from core.job_costing import JobCostingInput, calculate_selling_price
    jc = calculate_selling_price(JobCostingInput(
        total_cabinet_count = max(total_cabs, 1),
        material_cost_usd   = total_material,
        gp_target_pct       = config.get("gp_target_pct",   0.35),
        commission_pct      = config.get("commission_pct",  0.05),
        bond_pct            = config.get("bond_pct",        0.015),
    ))

    costing = {
        "material_cost":       jc.material_cost,
        "local_use_tax":       jc.local_use_tax,
        "ocean_freight":       jc.ocean_freight,
        "inland_delivery":     jc.inland_delivery,
        "installation":        jc.installation,
        "warehousing":         jc.warehousing,
        "material_protection": jc.material_protection,
        "insurance":           jc.insurance,
        "misc_allowance":      jc.misc_allowance,
        "pre_margin_total":    jc.pre_margin_total,
        "selling_price":       jc.selling_price,
        "gross_profit":        jc.gross_profit,
        "total_cost":          jc.total_cost,
        "containers_needed":   jc.containers_needed,
        "cost_per_cabinet":    jc.cost_per_cabinet,
        "sell_per_cabinet":    jc.sell_per_cabinet,
        "gp_pct":              jc.gp_pct,
        "commission_pct":      jc.commission_pct,
        "bond_pct":            jc.bond_pct,
        "total_cabinet_count": total_cabs,
    }

    api_calls = sum(
        ev.get("api_calls", 1)
        for sched in unit_schedules.values()
        for ev in sched.get("elevations", [])
    )
    tok_in  = api_calls * 5_000
    tok_out = api_calls * 350
    metrics = {
        "api_calls":      api_calls,
        "input_tokens":   tok_in,
        "output_tokens":  tok_out,
        "total_cost_usd": tok_in / 1_000_000 * 1.5 + tok_out / 1_000_000 * 3.0,
        "primary_model":  "google/gemini-2.5-flash",
        "fallback_model": "google/gemini-2.5-pro",
    }

    return {
        "unit_schedules":      unit_schedules,
        "cabinet_line_items":  line_items,
        "costing":             costing,
        "config":              config,
        "metrics":             metrics,
        "validation":          validation,
        "outputs": {
            "pdf_ready":  (out_dir / f"{project_id}_Shop_Drawings.pdf").exists(),
            "xlsx_ready": (out_dir / f"{project_id}_Cabinet_Estimation.xlsx").exists(),
        },
    }


# ─────────────────────────────────────────────────────────────────────────────
# DOWNLOADS
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/projects/{project_id}/download/pdf")
async def download_pdf(project_id: str):
    p = OUTPUTS_ROOT / project_id / f"{project_id}_Shop_Drawings.pdf"
    if not p.exists():
        raise HTTPException(404, "PDF not generated yet — run the pipeline.")
    return FileResponse(str(p), filename=p.name, media_type="application/pdf")


@app.get("/api/projects/{project_id}/download/excel")
async def download_excel(project_id: str):
    p = OUTPUTS_ROOT / project_id / f"{project_id}_Cabinet_Estimation.xlsx"
    if not p.exists():
        raise HTTPException(404, "Excel not generated yet — run the pipeline.")
    return FileResponse(
        str(p), filename=p.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ─────────────────────────────────────────────────────────────────────────────
# CROPS + REGION OVERLAY + PDF THUMBNAIL
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/projects/{project_id}/crops/{filename}")
async def serve_crop(project_id: str, filename: str):
    p = OUTPUTS_ROOT / project_id / "crops" / filename
    if not p.exists():
        raise HTTPException(404, f"Crop not found: {filename}")
    return FileResponse(str(p), media_type="image/png")


@app.get("/api/projects/{project_id}/regions/{unit_type:path}")
async def get_regions(project_id: str, unit_type: str):
    result = _find_project(project_id)
    if result is None:
        raise HTTPException(404, "Config not found")
    config, _, _ = result

    unit_type = unit_type.replace("+", " ")
    unit_plan_pdfs = config.get("unit_plan_pdfs", {})
    if unit_type not in unit_plan_pdfs:
        raise HTTPException(404, f"Unit {unit_type!r} not in config")

    pdf_path = PROJECT_ROOT / unit_plan_pdfs[unit_type]
    if not pdf_path.exists():
        raise HTTPException(404, f"PDF not found: {pdf_path.name}")

    try:
        from core.pdf_extractor import PDFExtractor
        from core.region_detector import RegionDetector
        with PDFExtractor(pdf_path) as ex:
            page = ex.extract_page(0)
            regions = RegionDetector(
                page.spans, page.rects, page.page_w, page.page_h
            ).detect()
        return {
            "page_w":  page.page_w,
            "page_h":  page.page_h,
            "regions": [
                {
                    "region_type": r.region_type,
                    "label":       r.label,
                    "bbox":        [r.x0, r.y0, r.x1, r.y1],
                    "confidence":  r.confidence,
                }
                for r in regions
            ],
        }
    except Exception as exc:
        raise HTTPException(500, f"Region detection failed: {exc}")


@app.get("/api/projects/{project_id}/pdf-pages/{unit_type:path}")
async def pdf_page_image(project_id: str, unit_type: str):
    result = _find_project(project_id)
    if result is None:
        raise HTTPException(404, "Config not found")
    config, _, _ = result

    unit_type = unit_type.replace("+", " ")
    pdf_path  = PROJECT_ROOT / config.get("unit_plan_pdfs", {}).get(unit_type, "")
    if not pdf_path.exists():
        raise HTTPException(404, f"PDF not found for unit {unit_type!r}")

    try:
        doc = fitz.open(str(pdf_path))
        pix = doc[0].get_pixmap(dpi=120)
        img = pix.tobytes("png")
        doc.close()
        return Response(content=img, media_type="image/png")
    except Exception as exc:
        raise HTTPException(500, f"Render failed: {exc}")


# ─────────────────────────────────────────────────────────────────────────────
# STATIC FILES — must be mounted last
# ─────────────────────────────────────────────────────────────────────────────

app.mount(
    "/",
    StaticFiles(directory=str(_HERE / "static"), html=True),
    name="static",
)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("web_app.app:app", host="127.0.0.1", port=8080, reload=True,
                app_dir=str(PROJECT_ROOT))
