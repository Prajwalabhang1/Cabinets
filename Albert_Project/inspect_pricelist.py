import openpyxl

# Inspect Price List
xl_path = r"Casa familia\02_Price_List\MS PRICE LIST LEVEL 1 -90CM.xlsx"
wb = openpyxl.load_workbook(xl_path, data_only=True)
ws = wb.active

print(f"FILE: MS PRICE LIST LEVEL 1 -90CM.xlsx")
print(f"Sheet: {ws.title}")
print(f"Rows: {ws.max_row}   Cols: {ws.max_column}")
print()
print("--- ALL ROWS IN PRICE LIST ---")
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    if any(cell is not None for cell in row):
        print(f"  {row}")
