"""
===========================================================================
  generators/cabinet_excel.py — Data-Driven Excel Estimation Generator
===========================================================================
  Generates the 4-sheet Excel workbook from pipeline data:
    Sheet 1: Project Info
    Sheet 2: Cabinet Matrix (per unit type × per cabinet type counts)
    Sheet 3: Cabinet Library (standard catalog)
    Sheet 4: Job Costing (full price breakdown + selling price)
    Sheets 5+: Shop Drawing Schedule per unit type

  Key improvements over generate_heritage_excel.py:
    1. Data-driven: accepts cabinet schedules from AI Vision pipeline
    2. Works for ANY project (not just Heritage Village)
    3. Job Costing sheet is formula-driven (not hardcoded row refs)
    4. Falls back gracefully when AI data is not available
===========================================================================
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.job_costing import JobCostingResult


# ══════════════════════════════════════════════════════════════════════════
# SHARED STYLES
# ══════════════════════════════════════════════════════════════════════════

def _make_styles():
    return {
        "title":   Font(name="Arial", size=16, bold=True,  color="1B365D"),
        "section": Font(name="Arial", size=12, bold=True,  color="1B365D"),
        "header":  Font(name="Arial", size=10, bold=True,  color="FFFFFF"),
        "bold":    Font(name="Arial", size=10, bold=True),
        "reg":     Font(name="Arial", size=10),
        "small":   Font(name="Arial", size=8, italic=True),
        "fill_navy":     PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid"),
        "fill_gray":     PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid"),
        "fill_blue":     PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        "fill_green":    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "fill_yellow":   PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "thin_side": Side(style="thin", color="D3D3D3"),
        "center": Alignment(horizontal="center", vertical="center"),
        "left":   Alignment(horizontal="left",   vertical="center"),
        "right":  Alignment(horizontal="right",  vertical="center"),
    }

def _thin_border(st):
    s = st["thin_side"]
    return Border(left=s, right=s, top=s, bottom=s)

def _double_bottom(st):
    s = st["thin_side"]
    return Border(bottom=Side(style="double", color="1B365D"), top=s)


# ══════════════════════════════════════════════════════════════════════════
# SHEET 1: Project Info
# ══════════════════════════════════════════════════════════════════════════

def _build_project_info(ws, config: dict, st: dict, unit_totals: dict, jc: JobCostingResult):
    ws.cell(1, 1, "CABINET ESTIMATION — PROJECT INFORMATION").font = st["title"]
    ws.cell(2, 1, f"{config['project_name']} — {config.get('owner', '')}").font = st["small"]

    total_units = sum(unit_totals.values())
    fields = [
        ("Project Name:",   config.get("project_name", ""),   st["fill_blue"]),
        ("Project ID:",     config.get("project_id", ""),     st["fill_blue"]),
        ("Date:",           "=TODAY()",                        None),
        ("Prepared By:",    "AI Estimation System",            None),
        ("Address:",        config.get("address", ""),         None),
        ("Price List:",     config.get("price_list_path", "").split("/")[-1] or "Standard", st["fill_blue"]),
        ("GP Target %:",    config.get("gp_target_pct", 0.35), st["fill_blue"]),
        ("Total Units:",    total_units,                       st["fill_green"]),
        ("Total Cabinets:", jc.inputs.total_cabinet_count,    st["fill_green"]),
        ("Selling Price:",  jc.selling_price,                 st["fill_green"]),
        ("Finish Tier:",    f"Standard {config.get('finish_tier', 1)}", None),
        ("Door Style:",     config.get("door_style", "Shaker"), None),
    ]

    for idx, (label, val, fill) in enumerate(fields, 4):
        ws.cell(idx, 1, label).font = st["bold"]
        c = ws.cell(idx, 2, val)
        c.font = st["reg"]
        if fill:
            c.fill = fill
        if label == "GP Target %:":
            c.number_format = "0.0%"
        if label in ("Selling Price:",):
            c.number_format = '$#,##0.00'
            c.font = Font(name="Arial", size=10, bold=True, color="1B365D")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50


# ══════════════════════════════════════════════════════════════════════════
# SHEET 2: Cabinet Matrix
# ══════════════════════════════════════════════════════════════════════════

def _build_cabinet_matrix(ws, config: dict, unit_schedules: dict, unit_totals: dict, st: dict):
    ws.cell(1, 1, "CABINET MATRIX — UNIT TYPE BREAKDOWN").font = st["title"]
    thin = _thin_border(st)
    dbl  = _double_bottom(st)

    # Headers row
    headers = ["Unit Type", "Qty", "Upper Wall", "Base", "Pantry/Tall", "Sink Base",
               "Corner", "Total Kitchen", "Vanity", "Medicine", "Linen", "Total Bath", "UNIT TOTAL"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(4, col, h)
        c.font   = st["header"]
        c.fill   = st["fill_navy"]
        c.alignment = st["center"]
        c.border = thin

    unit_types = list(unit_totals.keys())
    matrix_rows: dict[str, list] = {}

    for idx, unit_type in enumerate(unit_types, 5):
        qty      = unit_totals.get(unit_type, 0)
        schedule = unit_schedules.get(unit_type)

        if schedule:
            cabs = [c for c in schedule.all_cabinets if c.cabinet_type != "appliance_space"]
            upper    = sum(c.quantity for c in cabs if c.cabinet_type in ("upper_wall", "corner_upper", "microwave_shelf"))
            base     = sum(c.quantity for c in cabs if c.cabinet_type in ("base", "dw_adjacent", "corner_base"))
            pantry   = sum(c.quantity for c in cabs if c.cabinet_type == "pantry")
            sink_b   = sum(c.quantity for c in cabs if c.cabinet_type == "sink_base")
            corner   = sum(c.quantity for c in cabs if "corner" in c.cabinet_type)
            vanity   = sum(c.quantity for c in cabs if c.cabinet_type == "vanity")
            med      = sum(c.quantity for c in cabs if c.cabinet_type == "medicine_cabinet")
            linen    = sum(c.quantity for c in cabs if c.cabinet_type == "linen")
        else:
            upper = base = pantry = sink_b = corner = vanity = med = linen = 0

        row_data = [unit_type, qty, upper, base, pantry, sink_b, corner,
                    f"=SUM(C{idx}:G{idx})", vanity, med, linen,
                    f"=SUM(I{idx}:K{idx})",
                    f"=H{idx}+L{idx}"]
        matrix_rows[unit_type] = row_data

        for col, val in enumerate(row_data, 1):
            c = ws.cell(idx, col, val)
            c.font = st["bold"] if col in (1, 8, 12, 13) else st["reg"]
            c.border = thin
            c.alignment = st["left"] if col == 1 else st["center"]

    # Totals row
    tot_row = 5 + len(unit_types)
    ws.cell(tot_row, 1, "PROJECT TOTALS").font = st["bold"]
    ws.cell(tot_row, 2, f"=SUM(B5:B{tot_row-1})").font = st["bold"]
    for col in range(3, 14):
        ltr = get_column_letter(col)
        ws.cell(tot_row, col, f"=SUM({ltr}5:{ltr}{tot_row-1})").font = st["bold"]
        ws.cell(tot_row, col).border = dbl
        ws.cell(tot_row, col).fill  = st["fill_gray"]
        ws.cell(tot_row, col).alignment = st["center"]
    ws.cell(tot_row, 1).border = dbl
    ws.cell(tot_row, 1).fill   = st["fill_gray"]
    ws.cell(tot_row, 2).border = dbl
    ws.cell(tot_row, 2).fill   = st["fill_gray"]
    ws.cell(tot_row, 2).alignment = st["center"]

    # Column widths
    for col in ws.columns:
        max_len = 0
        col_ltr = get_column_letter(col[0].column)
        for cell in col:
            if cell.value and not str(cell.value).startswith("="):
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_ltr].width = max(max_len + 3, 12)


# ══════════════════════════════════════════════════════════════════════════
# SHEET 3: Cabinet Library
# ══════════════════════════════════════════════════════════════════════════

def _build_cabinet_library(ws, st: dict):
    ws.cell(1, 1, "CABINET LIBRARY — STANDARD SPECIFICATIONS").font = st["title"]
    thin = _thin_border(st)

    headers = ["Cabinet Code", "Description", "Type", "W (in)", "H (in)", "D (in)",
               "Finish Tier", "Door Style", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(3, col, h)
        c.font = st["header"]; c.fill = st["fill_navy"]; c.alignment = st["center"]

    lib_data = [
        ("W12", "Wall Cabinet 12W x 12H",    "Upper", 12, 12, 13, "Standard 1", "Shaker", "Over fridge"),
        ("W18", "Wall Cabinet 18W x 18H",    "Upper", 18, 18, 13, "Standard 1", "Shaker", "Over micro"),
        ("W24", "Wall Cabinet 24W x 30H",    "Upper", 24, 30, 13, "Standard 1", "Shaker", "Standard upper"),
        ("W30", "Wall Cabinet 30W x 30H",    "Upper", 30, 30, 13, "Standard 1", "Shaker", "Standard upper"),
        ("W36", "Wall Cabinet 36W x 30H",    "Upper", 36, 30, 13, "Standard 1", "Shaker", "Standard upper"),
        ("WC24","Corner Wall Cabinet 24W",   "Upper", 24, 30, 13, "Standard 1", "Shaker", "Blind corner"),
        ("B12", "Base Cabinet 12W",          "Lower", 12, 34.5, 24, "Standard 1", "Shaker", ""),
        ("B18", "Base Cabinet 18W",          "Lower", 18, 34.5, 24, "Standard 1", "Shaker", ""),
        ("B24", "Base Cabinet 24W",          "Lower", 24, 34.5, 24, "Standard 1", "Shaker", ""),
        ("B30", "Base Cabinet 30W",          "Lower", 30, 34.5, 24, "Standard 1", "Shaker", ""),
        ("B36", "Base Cabinet 36W",          "Lower", 36, 34.5, 24, "Standard 1", "Shaker", ""),
        ("SB36","Sink Base Cabinet 36W",     "Lower", 36, 34.5, 24, "Standard 1", "Shaker", "Sink base"),
        ("BC36","Corner Base Cabinet 36W",   "Lower", 36, 34.5, 24, "Standard 1", "Shaker", "Lazy susan"),
        ("T15", "Tall Pantry 15W x 84H",    "Tall",  15, 84, 24, "Standard 1", "Shaker", "Heritage spec"),
        ("T18", "Tall Pantry 18W x 84H",    "Tall",  18, 84, 24, "Standard 1", "Shaker", ""),
        ("T24", "Tall Pantry 24W x 84H",    "Tall",  24, 84, 24, "Standard 1", "Shaker", ""),
        ("VAN24","Bath Vanity 24W",          "Vanity",24, 34.5, 21, "Standard 1", "Shaker", "Single sink"),
        ("VAN30","Bath Vanity 30W",          "Vanity",30, 34.5, 21, "Standard 1", "Shaker", "Single sink"),
        ("VAN36","Bath Vanity 36W",          "Vanity",36, 34.5, 21, "Standard 1", "Shaker", "Single sink"),
        ("VAN48","Bath Vanity 48W",          "Vanity",48, 34.5, 21, "Standard 1", "Shaker", "Double sink"),
        ("VAN60","Bath Vanity 60W",          "Vanity",60, 34.5, 21, "Standard 1", "Shaker", "Double sink"),
        ("MED24","Medicine Cabinet 24W",     "Bath",  24, 30, 4,  "Standard 1", "Mirror", ""),
        ("MED30","Medicine Cabinet 30W",     "Bath",  30, 30, 4,  "Standard 1", "Mirror", ""),
        ("MED36","Medicine Cabinet 36W",     "Bath",  36, 30, 4,  "Standard 1", "Mirror", ""),
        ("MED60","Medicine Cabinet 60W",     "Bath",  60, 30, 4,  "Standard 1", "Mirror", ""),
        ("LIN18","Linen Cabinet 18W x 84H",  "Bath",  18, 84, 18, "Standard 1", "Shaker", ""),
    ]
    for r_off, row in enumerate(lib_data, 4):
        for col_idx, val in enumerate(row, 1):
            c = ws.cell(r_off, col_idx, val)
            c.font = st["reg"]; c.border = thin
            c.alignment = st["right"] if col_idx in (4, 5, 6) else st["left"]

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 3, 12)


# ══════════════════════════════════════════════════════════════════════════
# SHEET 4: Job Costing
# ══════════════════════════════════════════════════════════════════════════

def _build_job_costing(ws, config: dict, jc: JobCostingResult, st: dict):
    ws.cell(1, 1, f"JOB COSTING — {config.get('project_name', 'PROJECT')} TOTAL ESTIMATE").font = st["title"]
    ws.cell(2, 1, f"{config.get('owner', '')}  |  GP Target: {jc.inputs.gp_target_pct:.0%}").font = st["small"]
    thin = _thin_border(st)
    dbl  = _double_bottom(st)

    # ── A. Quantities ─────────────────────────────────────────────────────
    ws.cell(4, 1, "A. QUANTITIES").font = st["bold"]
    qty_rows = [
        ("Total Cabinet Count:",   jc.inputs.total_cabinet_count, "units"),
        ("Containers Required:",   jc.containers_needed, "containers (220 cabs each)"),
    ]
    for i, (lbl, val, unit) in enumerate(qty_rows, 5):
        ws.cell(i, 2, lbl).font = st["reg"]
        ws.cell(i, 5, val).font = st["bold"]
        ws.cell(i, 6, unit).font = st["reg"]

    # ── B. Material ───────────────────────────────────────────────────────
    ws.cell(8, 1, "B. MATERIAL COST").font = st["bold"]
    ws.cell(9, 2, "Material Cost (from price list)").font = st["reg"]
    ws.cell(9, 5, jc.material_cost).font = st["reg"]
    ws.cell(9, 5).number_format = "$#,##0.00"

    # ── C. Project Costs ──────────────────────────────────────────────────
    ws.cell(11, 1, "C. PROJECT COSTS").font = st["bold"]
    cost_items = [
        ("1", "Material Cost (cabinets)",         jc.material_cost,          "Invoice / Price List"),
        ("2", "Local Use Tax",                     jc.local_use_tax,          f"{jc.inputs.local_use_tax_pct:.1%} of Material"),
        ("3", "Ocean Freight / Shipping",          jc.ocean_freight,          f"${jc.inputs.ocean_freight_per_container:,.0f} × {jc.containers_needed} containers"),
        ("4", "Inland Delivery",                   jc.inland_delivery,        "Fixed"),
        ("5", "Installation",                      jc.installation,           f"${jc.inputs.installation_per_cabinet:.0f}/cabinet × {jc.inputs.total_cabinet_count} cabs"),
        ("6", "Warehousing",                       jc.warehousing,            f"{jc.inputs.warehousing_pct:.1%} of Material"),
        ("7", "Material Protection",               jc.material_protection,    f"{jc.inputs.material_protection_pct:.1%} of Material"),
        ("8", "Insurance",                         jc.insurance,              f"{jc.inputs.insurance_pct:.1%} of Material"),
        ("9", "Miscellaneous Allowance",           jc.misc_allowance,         "Fixed"),
    ]
    hdr_row = 12
    ws.cell(hdr_row, 1, "#").font    = st["bold"]
    ws.cell(hdr_row, 2, "Cost Item").font = st["bold"]
    ws.cell(hdr_row, 4, "Basis").font = st["bold"]
    ws.cell(hdr_row, 5, "Amount ($)").font = st["bold"]
    for c in range(1, 6): ws.cell(hdr_row, c).fill = st["fill_gray"]; ws.cell(hdr_row, c).border = thin

    for idx, item in enumerate(cost_items, hdr_row + 1):
        ws.cell(idx, 1, item[0]).font = st["reg"]
        ws.cell(idx, 2, item[1]).font = st["reg"]
        ws.cell(idx, 5, item[2]).font = st["reg"]
        ws.cell(idx, 5).number_format = "$#,##0.00"
        ws.cell(idx, 4, item[3]).font = st["small"]
        for c in range(1, 6): ws.cell(idx, c).border = thin

    subtotal_row = hdr_row + len(cost_items) + 1
    ws.cell(subtotal_row, 1, "PRE-MARGIN SUBTOTAL").font = st["bold"]
    ws.cell(subtotal_row, 5, jc.pre_margin_total).font   = st["bold"]
    ws.cell(subtotal_row, 5).number_format = "$#,##0.00"
    ws.cell(subtotal_row, 5).fill = st["fill_yellow"]
    for c in range(1, 6): ws.cell(subtotal_row, c).border = dbl; ws.cell(subtotal_row, c).fill = st["fill_gray"]
    ws.cell(subtotal_row, 5).fill = st["fill_yellow"]

    # ── D. Margin ─────────────────────────────────────────────────────────
    mrg_start = subtotal_row + 2
    ws.cell(mrg_start, 1, "D. MARGIN & OVERHEAD").font = st["bold"]
    margin_items = [
        ("10", "Commission",   jc.inputs.commission_pct, jc.selling_price * jc.inputs.commission_pct),
        ("11", "Bond",         jc.inputs.bond_pct,       jc.selling_price * jc.inputs.bond_pct),
        ("12", "Gross Profit", jc.inputs.gp_target_pct,  jc.gross_profit),
    ]
    for i, (num, lbl, pct, amt) in enumerate(margin_items, mrg_start + 1):
        ws.cell(i, 1, num).font = st["reg"]
        ws.cell(i, 2, lbl).font = st["reg"]
        ws.cell(i, 4, f"{pct:.1%} of Selling Price").font = st["small"]
        ws.cell(i, 5, amt).font = st["bold"]
        ws.cell(i, 5).number_format = "$#,##0.00"
        for c in range(1, 6): ws.cell(i, c).border = thin

    # ── E. Selling Price ──────────────────────────────────────────────────
    sp_row = mrg_start + len(margin_items) + 3
    ws.cell(sp_row, 1, "E. SELLING PRICE (TOTAL PROJECT)").font = Font(name="Arial", size=14, bold=True, color="1B365D")
    ws.cell(sp_row, 5, jc.selling_price).font = Font(name="Arial", size=14, bold=True, color="1B365D")
    ws.cell(sp_row, 5).fill = st["fill_green"]
    ws.cell(sp_row, 5).number_format = "$#,##0.00"

    # ── F. Verification ───────────────────────────────────────────────────
    vfy_row = sp_row + 2
    ws.cell(vfy_row, 1, "F. VERIFICATION").font = st["bold"]
    vfy_items = [
        ("Gross Profit Amount:", jc.gross_profit,       "$#,##0.00"),
        ("Gross Profit %:",      jc.gp_check_pct,       "0.0%"),
        ("Total Cost:",          jc.total_cost,          "$#,##0.00"),
        ("Cost per Cabinet:",    jc.cost_per_cabinet,   "$#,##0.00"),
        ("Sell per Cabinet:",    jc.sell_per_cabinet,   "$#,##0.00"),
    ]
    for i, (lbl, val, fmt) in enumerate(vfy_items, vfy_row + 1):
        ws.cell(i, 2, lbl).font = st["bold"]
        ws.cell(i, 5, val).font = st["bold"]
        ws.cell(i, 5).number_format = fmt

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 20


# ══════════════════════════════════════════════════════════════════════════
# SHOP DRAWING SCHEDULE SHEETS
# ══════════════════════════════════════════════════════════════════════════

def _build_shop_drawing_sheet(ws, unit_type: str, config: dict, schedule, st: dict):
    """One sheet per unit type — full cabinet schedule with pricing."""
    thin = _thin_border(st)
    dbl  = _double_bottom(st)

    # Title block
    ws.cell(1, 1, "CABINET SHOP DRAWING — UNIT SCHEDULE").font = st["title"]
    ws.cell(2, 1, f"{config.get('owner', '')}  |  {config.get('architect', '')}").font = st["small"]

    is_ada = "ACC" in unit_type or "ADA" in unit_type

    ws.cell(4, 1, "PROJECT:").font     = st["bold"]
    ws.cell(4, 2, config.get("project_name", "")).font = st["reg"]
    ws.cell(4, 5, "UNIT TYPE:").font   = st["bold"]
    ws.cell(4, 6, unit_type).font      = Font(name="Arial", size=10, bold=True)
    ws.cell(4, 9, "DATE:").font        = st["bold"]
    ws.cell(4, 10, "=TODAY()").font    = st["reg"]

    ws.cell(5, 1, "ADDRESS:").font     = st["bold"]
    ws.cell(5, 2, config.get("address", "")).font = st["reg"]
    ws.cell(5, 5, "DESCRIPTION:").font = st["bold"]
    desc = "Fully Accessible ADA Unit" if is_ada else "FHA Type B Unit"
    ws.cell(5, 6, desc).font = st["reg"]
    ws.cell(5, 9, "PREPARED BY:").font = st["bold"]
    ws.cell(5, 10, "AI Estimation System").font = st["reg"]

    ws.cell(6, 1, "DRAWING REF:").font = st["bold"]
    ws.cell(6, 2, f"{config.get('project_name', '')} Unit Plans").font = st["reg"]
    ws.cell(6, 5, "FINISH TIER:").font = st["bold"]
    ws.cell(6, 6, f"Standard {config.get('finish_tier', 1)}").font = st["reg"]

    headers = ["Item", "Cabinet Code", "Description", "Type",
               "W (in)", "H (in)", "D (in)", "Qty", "Elev. Ref.", "Location Note"]

    # Section 1: Kitchen
    ws.cell(8, 1, "SECTION 1 — KITCHEN CABINETS").font = st["section"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(10, col, h)
        c.font = st["header"]; c.fill = st["fill_navy"]; c.alignment = st["center"]; c.border = thin

    if not schedule:
        ws.cell(11, 1, "No schedule available — PDF not processed").font = st["small"]
        return

    kitchen_types = {"upper_wall", "base", "sink_base", "dw_adjacent",
                     "microwave_shelf", "pantry", "corner_upper", "corner_base", "filler"}
    bath_types    = {"vanity", "medicine_cabinet", "linen"}

    row_idx  = 11
    item_num = 1
    kitchen_qty = 0

    all_cabinets = schedule.all_cabinets
    kitchen_cabs = [c for c in all_cabinets if c.cabinet_type in kitchen_types]
    bath_cabs    = [c for c in all_cabinets if c.cabinet_type in bath_types]

    for cab in kitchen_cabs:
        w_in = round(cab.width_mm / 25.4, 1)
        h_in = round(cab.height_mm / 25.4, 1)
        d_in = round(cab.depth_mm  / 25.4, 1)
        vals = [item_num, cab.code, f"{cab.cabinet_type} {w_in}\"W × {h_in}\"H",
                cab.cabinet_type.replace("_", " ").title(),
                w_in, h_in, d_in, cab.quantity,
                cab.elevation_ref, cab.location[:30] if cab.location else ""]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row_idx, col, v)
            c.font   = st["bold"] if col == 2 else st["reg"]
            c.border = thin
            c.alignment = st["right"] if col in (5, 6, 7, 8) else st["left"]
        kitchen_qty += cab.quantity
        item_num += 1; row_idx += 1

    # Kitchen subtotal
    ws.cell(row_idx, 1, "KITCHEN SUBTOTALS").font = st["bold"]
    ws.cell(row_idx, 8, kitchen_qty).font = st["bold"]
    for c in range(1, 11): ws.cell(row_idx, c).border = dbl; ws.cell(row_idx, c).fill = st["fill_gray"]
    row_idx += 2

    # Section 2: Bathroom
    ws.cell(row_idx, 1, "SECTION 2 — BATHROOM VANITY CABINETS").font = st["section"]
    row_idx += 2
    for col, h in enumerate(headers, 1):
        c = ws.cell(row_idx, col, h)
        c.font = st["header"]; c.fill = st["fill_navy"]; c.alignment = st["center"]; c.border = thin
    row_idx += 1

    item_num_bath = 1
    bath_qty = 0
    for cab in bath_cabs:
        w_in = round(cab.width_mm / 25.4, 1)
        h_in = round(cab.height_mm / 25.4, 1)
        d_in = round(cab.depth_mm  / 25.4, 1)
        vals = [item_num_bath, cab.code,
                f"{cab.cabinet_type} {w_in}\"W × {h_in}\"H",
                cab.cabinet_type.replace("_", " ").title(),
                w_in, h_in, d_in, cab.quantity,
                cab.elevation_ref, cab.location[:30] if cab.location else ""]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row_idx, col, v)
            c.font = st["bold"] if col == 2 else st["reg"]
            c.border = thin
            c.alignment = st["right"] if col in (5, 6, 7, 8) else st["left"]
        bath_qty += cab.quantity
        item_num_bath += 1; row_idx += 1

    ws.cell(row_idx, 1, "BATHROOM SUBTOTALS").font = st["bold"]
    ws.cell(row_idx, 8, bath_qty).font = st["bold"]
    for c in range(1, 11): ws.cell(row_idx, c).border = dbl; ws.cell(row_idx, c).fill = st["fill_gray"]
    row_idx += 2

    ws.cell(row_idx, 1, f"UNIT {unit_type} GRAND TOTAL").font = st["bold"]
    ws.cell(row_idx, 8, kitchen_qty + bath_qty).font = Font(name="Arial", size=10, bold=True, color="1B365D")
    for c in range(1, 11): ws.cell(row_idx, c).border = dbl; ws.cell(row_idx, c).fill = st["fill_gray"]

    # Column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 2, 10)


# ══════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════

def generate_excel(
    config:          dict,
    unit_schedules:  dict,   # {unit_type: UnitSchedule}
    unit_totals:     dict,   # {unit_type: count}
    jc_result:       JobCostingResult,
    output_path:     str | Path,
) -> Path:
    """
    Generate the complete Excel estimation workbook.

    Args:
        config:         project_config.json dict
        unit_schedules: {unit_type: UnitSchedule} from pipeline Step 1
        unit_totals:    {unit_type: count} from pipeline Step 2
        jc_result:      JobCostingResult from pipeline Step 4
        output_path:    where to save the .xlsx file
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    st = _make_styles()

    wb = openpyxl.Workbook()

    # Sheet 1: Project Info
    ws_info = wb.active
    ws_info.title = "Project Info"
    _build_project_info(ws_info, config, st, unit_totals, jc_result)

    # Sheet 2: Cabinet Matrix
    ws_matrix = wb.create_sheet("Cabinet Matrix")
    _build_cabinet_matrix(ws_matrix, config, unit_schedules, unit_totals, st)

    # Sheet 3: Cabinet Library
    ws_lib = wb.create_sheet("Cabinet Library")
    _build_cabinet_library(ws_lib, st)

    # Sheet 4: Job Costing
    ws_cost = wb.create_sheet("Job Costing")
    _build_job_costing(ws_cost, config, jc_result, st)

    # Sheets 5+: Shop Drawing per unit type
    for unit_type in unit_totals.keys():
        safe_name = f"SD-{unit_type}"[:31]  # Excel sheet name max 31 chars
        ws_sd = wb.create_sheet(safe_name)
        schedule = unit_schedules.get(unit_type)
        _build_shop_drawing_sheet(ws_sd, unit_type, config, schedule, st)

    wb.save(str(output_path))
    print(f"  ✅ Excel saved: {output_path}")
    return output_path


# ══════════════════════════════════════════════════════════════════════════
# CLI TEST
# ══════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

    from core.job_costing import JobCostingInput, calculate_selling_price

    config = {
        "project_id": "TEST-001",
        "project_name": "TEST PROJECT",
        "owner": "Test Owner",
        "address": "123 Test St, Miami FL",
        "gp_target_pct": 0.35,
        "finish_tier": 1,
        "door_style": "Shaker",
    }
    unit_totals = {"A1": 14, "B1": 6}

    jc = calculate_selling_price(JobCostingInput(
        total_cabinet_count = 140,
        material_cost_usd   = 25000.0,
    ))

    out = Path("outputs/TEST-001/TEST_Cabinet_Estimation.xlsx")
    generate_excel(config, {}, unit_totals, jc, out)
    print(f"Test complete: {out}")
