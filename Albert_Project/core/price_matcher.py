"""
===========================================================================
  core/price_matcher.py — Euro Price List → USD Cabinet Matcher
===========================================================================
  Reads the MS PRICE LIST LEVEL 1-90CM.xlsx vendor catalog and matches
  extracted cabinet codes/dimensions to their EUR prices, converting to USD.

  Key challenges:
    - Price list uses mm (metric), drawings use inches → conversion needed
    - Price list cabinet types don't map 1:1 to architect labels
    - Standard 1 / Standard 2 / Standard 3 tiers (different price columns)
    - Fuzzy matching needed (AI-extracted widths may differ by ±25mm from list)

  Price list structure (actual file):
    Col B: Description (e.g., "Base Cabinet 1 Door")
    Col C: Width (mm)
    Col D: Height (mm)
    Col E: Depth (mm)
    Col F: Price Standard 1 (EUR)
    Col G: Price Standard 2 (EUR)  [if present]
    Col H: Price Standard 3 (EUR)  [if present]
===========================================================================
"""
from __future__ import annotations

import math
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from core.config import EUR_USD_RATE, CABINET_WIDTH_TOLERANCE_MM


# ══════════════════════════════════════════════════════════════════════════
# DATA CLASSES
# ══════════════════════════════════════════════════════════════════════════

@dataclass
class PriceListEntry:
    """One line item from the vendor price list."""
    description: str
    cabinet_type: str    # mapped type: "base" | "upper_wall" | "pantry" | "vanity" | ...
    width_mm:  float
    height_mm: float
    depth_mm:  float
    price_eur_tier1: float
    price_eur_tier2: Optional[float]
    price_eur_tier3: Optional[float]

    def price_usd(self, tier: int = 1, eur_usd: float = EUR_USD_RATE) -> float:
        """Return price in USD for the specified tier."""
        prices = {
            1: self.price_eur_tier1,
            2: self.price_eur_tier2 or self.price_eur_tier1,
            3: self.price_eur_tier3 or self.price_eur_tier2 or self.price_eur_tier1,
        }
        return (prices.get(tier, self.price_eur_tier1) or 0.0) * eur_usd


@dataclass
class MatchResult:
    """Result of matching a cabinet to the price list."""
    cabinet_code:    str
    cabinet_type:    str
    width_mm:        float
    matched_entry:   Optional[PriceListEntry]
    price_usd:       float
    price_eur:       float
    match_delta_mm:  float    # how far the matched width is from requested
    match_quality:   str      # "EXACT" | "FUZZY" | "FALLBACK" | "NOT_FOUND"
    quantity:        int = 1

    @property
    def total_usd(self) -> float:
        return self.price_usd * self.quantity

    @property
    def total_eur(self) -> float:
        return self.price_eur * self.quantity


# ══════════════════════════════════════════════════════════════════════════
# CABINET TYPE MAPPING
# ══════════════════════════════════════════════════════════════════════════

# Maps AI cabinet types → price list category keywords
# Order matters: more specific matches first
_TYPE_KEYWORD_MAP = {
    "sink_base":        ["sink base", "sink"],
    "dw_adjacent":      ["base", "1 door"],   # standard base adjacent to DW
    "corner_base":      ["corner base", "corner"],
    "corner_upper":     ["corner wall", "corner"],
    "microwave_shelf":  ["short wall", "wall"],
    "upper_wall":       ["medium wall", "tall wall", "wall cabinet", "wall"],
    "pantry":           ["tall pantry", "pantry"],
    "base":             ["base cabinet", "base"],
    "vanity":           ["vanity", "bathroom"],
    "medicine_cabinet": ["medicine", "mirror"],
    "linen":            ["linen", "tall"],
    "filler":           ["filler", "panel"],
}


def _map_type_to_price_category(cabinet_type: str) -> list[str]:
    """Return ordered list of keywords to search in price list descriptions."""
    return _TYPE_KEYWORD_MAP.get(cabinet_type, ["base"])


# ══════════════════════════════════════════════════════════════════════════
# PRICE LIST LOADER
# ══════════════════════════════════════════════════════════════════════════

class PriceMatcher:
    """
    Load vendor price list and match cabinets to prices.

    Usage:
        matcher = PriceMatcher("02_Price_List/MS PRICE LIST LEVEL 1 -90CM.xlsx")
        result  = matcher.match("base", width_mm=900, quantity=14, tier=1)
        print(f"${result.price_usd:.2f} per unit | ${result.total_usd:.2f} total")
    """

    def __init__(self, xlsx_path: str | Path, tier: int = 1, eur_usd_rate: float = EUR_USD_RATE):
        self.xlsx_path   = Path(xlsx_path)
        self.tier        = tier
        self.eur_usd     = eur_usd_rate
        self._entries:   list[PriceListEntry] = []
        self._loaded     = False

        if not self.xlsx_path.exists():
            raise FileNotFoundError(f"Price list not found: {self.xlsx_path}")

    def _ensure_loaded(self):
        if not self._loaded:
            self._load()
            self._loaded = True

    def _load(self):
        """Parse the Excel price list."""
        import openpyxl

        wb = openpyxl.load_workbook(str(self.xlsx_path), data_only=True)
        ws = wb.active

        print(f"  Loading price list: {self.xlsx_path.name}")
        print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")

        entries = []
        # Start from row 4 (rows 1-3 are typically headers)
        for row_num in range(4, ws.max_row + 1):
            desc  = ws.cell(row_num, 2).value   # Col B: Description
            w     = ws.cell(row_num, 3).value   # Col C: Width mm
            h     = ws.cell(row_num, 4).value   # Col D: Height mm
            d     = ws.cell(row_num, 5).value   # Col E: Depth mm
            p1    = ws.cell(row_num, 6).value   # Col F: Price tier 1
            p2    = ws.cell(row_num, 7).value   # Col G: Price tier 2
            p3    = ws.cell(row_num, 8).value   # Col H: Price tier 3

            if not desc or not w or not p1:
                continue

            try:
                desc_str = str(desc).strip()
                width_mm = float(w)
                height_mm = float(h) if h else 720.0
                depth_mm = float(d) if d else 600.0
                price1   = float(p1)
                price2   = float(p2) if p2 else None
                price3   = float(p3) if p3 else None
            except (ValueError, TypeError):
                continue

            # Map description to cabinet type
            cab_type = _infer_type_from_description(desc_str)

            entries.append(PriceListEntry(
                description      = desc_str,
                cabinet_type     = cab_type,
                width_mm         = width_mm,
                height_mm        = height_mm,
                depth_mm         = depth_mm,
                price_eur_tier1  = price1,
                price_eur_tier2  = price2,
                price_eur_tier3  = price3,
            ))

        self._entries = entries
        print(f"  Loaded {len(entries)} price list entries")

    @property
    def entries(self) -> list[PriceListEntry]:
        self._ensure_loaded()
        return self._entries

    # ── Main Matching Function ────────────────────────────────────────────

    def match(
        self,
        cabinet_type: str,
        width_mm:     float,
        quantity:     int   = 1,
        tier:         int   = None,
        tolerance_mm: float = CABINET_WIDTH_TOLERANCE_MM,
    ) -> MatchResult:
        """
        Find the best price list match for a cabinet.

        Args:
            cabinet_type: from VALID_CABINET_TYPES (e.g., "base", "upper_wall")
            width_mm:     cabinet width in mm
            quantity:     number of units
            tier:         price tier (1, 2, or 3). Defaults to self.tier.
            tolerance_mm: max width deviation for fuzzy match

        Returns:
            MatchResult with price and match quality
        """
        self._ensure_loaded()
        tier = tier or self.tier
        kw_list = _map_type_to_price_category(cabinet_type)

        best_entry:   Optional[PriceListEntry] = None
        best_delta:   float = math.inf
        match_quality = "NOT_FOUND"

        # Exact match first
        for entry in self._entries:
            desc_lower = entry.description.lower()
            type_match = any(kw in desc_lower for kw in kw_list)
            if not type_match:
                continue

            delta = abs(entry.width_mm - width_mm)
            if delta <= 1.0:  # exact
                best_entry   = entry
                best_delta   = delta
                match_quality = "EXACT"
                break
            elif delta <= tolerance_mm and delta < best_delta:
                best_entry   = entry
                best_delta   = delta
                match_quality = "FUZZY"

        # If no match, try broader search by width only
        if best_entry is None:
            for entry in self._entries:
                delta = abs(entry.width_mm - width_mm)
                if delta <= tolerance_mm * 1.5 and delta < best_delta:
                    best_entry   = entry
                    best_delta   = delta
                    match_quality = "FALLBACK"

        if best_entry is None:
            # Return zero price with NOT_FOUND flag
            return MatchResult(
                cabinet_code  = f"{cabinet_type.upper()}-{width_mm:.0f}mm",
                cabinet_type  = cabinet_type,
                width_mm      = width_mm,
                matched_entry = None,
                price_usd     = 0.0,
                price_eur     = 0.0,
                match_delta_mm = 0.0,
                match_quality = "NOT_FOUND",
                quantity      = quantity,
            )

        price_eur = best_entry.price_eur_tier1  # default
        if tier == 2 and best_entry.price_eur_tier2:
            price_eur = best_entry.price_eur_tier2
        elif tier == 3 and best_entry.price_eur_tier3:
            price_eur = best_entry.price_eur_tier3

        price_usd = (price_eur or 0.0) * self.eur_usd

        return MatchResult(
            cabinet_code  = _generate_code(cabinet_type, width_mm),
            cabinet_type  = cabinet_type,
            width_mm      = width_mm,
            matched_entry = best_entry,
            price_usd     = price_usd,
            price_eur     = price_eur or 0.0,
            match_delta_mm = best_delta,
            match_quality = match_quality,
            quantity      = quantity,
        )

    def price_schedule(
        self,
        cabinet_items: list,   # list of CabinetItem from ai_vision_classifier
        tier: int = None,
    ) -> list[MatchResult]:
        """Match an entire cabinet schedule at once."""
        tier = tier or self.tier
        results = []
        for item in cabinet_items:
            if item.cabinet_type == "appliance_space":
                continue  # no price for appliances
            result = self.match(
                cabinet_type = item.cabinet_type,
                width_mm     = item.width_mm,
                quantity     = item.quantity,
                tier         = tier,
            )
            results.append(result)
        return results

    def total_material_cost_usd(self, match_results: list[MatchResult]) -> float:
        """Sum total material cost from a list of match results."""
        return sum(r.total_usd for r in match_results)

    def print_pricing_report(self, match_results: list[MatchResult]):
        """Print a formatted pricing report."""
        print(f"\n  {'Cabinet Code':25s} {'Type':20s} {'W(mm)':8s} "
              f"{'Qty':5s} {'$/unit':10s} {'Total':12s} {'Match':12s}")
        print("  " + "-" * 100)
        total = 0.0
        not_found = 0
        for r in match_results:
            status = r.match_quality
            if r.match_quality == "NOT_FOUND":
                not_found += 1
            line = (f"  {r.cabinet_code:25s} {r.cabinet_type:20s} "
                    f"{r.width_mm:8.0f} {r.quantity:5d} "
                    f"${r.price_usd:9.2f} ${r.total_usd:11.2f}  [{status}]")
            print(line)
            total += r.total_usd
        print("  " + "-" * 100)
        print(f"  TOTAL MATERIAL COST (USD): ${total:,.2f}")
        if not_found:
            print(f"  ⚠️  {not_found} items NOT found in price list")
        return total


# ══════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════

def _infer_type_from_description(desc: str) -> str:
    """Map a price list description string to a cabinet type."""
    dl = desc.lower()
    if "sink" in dl:          return "sink_base"
    if "corner" in dl:
        if "wall" in dl:      return "corner_upper"
        return "corner_base"
    if "tall pantry" in dl or "pantry" in dl: return "pantry"
    if "tall wall" in dl:     return "upper_wall"
    if "medium wall" in dl:   return "upper_wall"
    if "short wall" in dl:    return "microwave_shelf"
    if "wall" in dl:          return "upper_wall"
    if "vanity" in dl:        return "vanity"
    if "medicine" in dl or "mirror" in dl: return "medicine_cabinet"
    if "linen" in dl:         return "linen"
    if "filler" in dl or "panel" in dl:   return "filler"
    if "base" in dl:          return "base"
    return "base"


def _generate_code(cabinet_type: str, width_mm: float) -> str:
    """Generate a short cabinet code from type and width."""
    w_in = round(width_mm / 25.4)
    prefixes = {
        "upper_wall":       f"W{w_in}",
        "base":             f"B{w_in}",
        "sink_base":        f"SB{w_in}",
        "dw_adjacent":      f"DWA{w_in}",
        "microwave_shelf":  f"MW{w_in}",
        "pantry":           f"T{w_in}",
        "corner_upper":     f"WC{w_in}",
        "corner_base":      f"BC{w_in}",
        "vanity":           f"VAN{w_in}",
        "medicine_cabinet": f"MED{w_in}",
        "linen":            f"LIN{w_in}",
        "filler":           f"FIL{w_in}",
    }
    return prefixes.get(cabinet_type, f"CAB{w_in}")


# ══════════════════════════════════════════════════════════════════════════
# FALLBACK: HARDCODED USD PRICES
# ══════════════════════════════════════════════════════════════════════════
# Used when price list Excel is unavailable or entry not found

FALLBACK_USD_CATALOG = {
    # Upper/Wall cabinets
    "W12":   89.0,  "W18":  95.0,  "W24": 105.0,  "W30": 115.0,  "W36": 129.0,
    # Base cabinets
    "B12":   79.0,  "B18":  89.0,  "B24": 105.0,  "B30": 119.0,  "B36": 139.0,
    # Special
    "BC36": 165.0,  "SB36": 159.0, "DWA24": 109.0,
    "T18":  199.0,  "T24": 229.0,
    # Vanity
    "VAN24": 189.0, "VAN30": 219.0, "VAN36": 249.0,
    "VAN48": 319.0, "VAN60": 379.0,
    # Bath
    "MED24":  79.0, "MED30":  99.0, "MED36": 119.0, "MED60": 179.0,
    "LIN18": 199.0,
}


def get_fallback_price_usd(cabinet_code: str) -> float:
    """Lookup fallback USD price by code. Returns 100.0 if not found."""
    return FALLBACK_USD_CATALOG.get(cabinet_code, 100.0)


# ══════════════════════════════════════════════════════════════════════════
# QUICK TEST
# ══════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python -m core.price_matcher <path_to_xlsx>")
        print("\nRunning fallback catalog test instead...")
        tests = [
            ("base",       900.0, 14),
            ("upper_wall", 762.0, 14),
            ("vanity",     900.0, 28),
            ("pantry",     380.0,  6),
        ]
        for cab_type, width, qty in tests:
            code  = _generate_code(cab_type, width)
            price = get_fallback_price_usd(code)
            total = price * qty
            print(f"  {code:10s}  ${price:8.2f}/unit × {qty:3d}  = ${total:10.2f}")
        sys.exit(0)

    xlsx_path = sys.argv[1]
    matcher = PriceMatcher(xlsx_path)

    tests = [
        ("base",       900.0, 14),
        ("upper_wall", 762.0, 14),
        ("sink_base",  900.0, 14),
        ("vanity",     900.0, 28),
        ("pantry",     380.0,  6),
        ("base",       300.0,  6),  # edge case
    ]

    results = []
    for cab_type, width, qty in tests:
        r = matcher.match(cab_type, width, quantity=qty)
        results.append(r)

    matcher.print_pricing_report(results)
